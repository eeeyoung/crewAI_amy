"""Client Variation Excel template engine.

Provides:
  - TemplateMapping: reads a YAML/JSON config that maps logical fields to Excel
    cell coordinates, making the system adaptable to different templates.
  - VariationExcelBuilder: opens a master template, fills cells, writes formulas,
    and saves the populated workbook.

All cell positions are derived from the mapping config — nothing is hardcoded.
"""
from __future__ import annotations

import copy
import json
import os
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color, numbers

import yaml


# =============================================================================
# Template Mapping — config → typed accessors
# =============================================================================

@dataclass
class TemplateMapping:
    """Reads a template mapping dict (from YAML/JSON) and provides typed accessors
    for cell positions. Decouples code from specific Excel layouts."""

    config: dict

    # ── VO Sheet ──────────────────────────────────────────────────────

    @property
    def blank_sheet_name(self) -> str:
        return self.config.get("template", {}).get("blank_sheet", "VOXX")

    @property
    def register_sheet_name(self) -> str:
        return self.config.get("template", {}).get("register_sheet", "Register")

    def vo_cell(self, field: str) -> str:
        """Return cell reference for a named VO sheet field.
        e.g., field='project_name' → 'B6'"""
        return self.config.get("vo_sheet", {}).get(field, "")

    @property
    def vo_items_start_row(self) -> int:
        return self.config.get("vo_sheet", {}).get("items_start_row", 14)

    @property
    def vo_items_max_rows(self) -> int:
        return self.config.get("vo_sheet", {}).get("items_max_rows", 5)

    @property
    def vo_items_end_row(self) -> int:
        return self.vo_items_start_row + self.vo_items_max_rows - 1

    def vo_item_col(self, field: str) -> str:
        """Return column letter for a line-item field.
        e.g., field='description' → 'B'"""
        return self.config.get("vo_sheet", {}).get("items_columns", {}).get(field, "")

    @property
    def vo_title_row(self) -> int | None:
        """Row number of the VO title cell (B11), or None."""
        cell = self.vo_cell("vo_title")
        if cell:
            return int(cell[1:])  # "B11" → 11
        return None

    # ── Totals Section (label-based detection) ────────────────────────

    @property
    def totals_labels(self) -> dict[str, str]:
        """Labels used to detect totals rows. e.g., {'sub_total': 'SUB TOTAL', ...}"""
        return self.config.get("vo_sheet", {}).get("totals_labels", {
            "sub_total": "SUB TOTAL",
            "nett_cost": "NETT VARIATION COST",
            "margin": "MARGIN AND OVERHEAD COSTS",
            "excl_gst": "VARIATION COST EXCLUDING GST",
            "gst": "GST",
            "total": "TOTAL INCLUDING GST",
            "raised_by": "Variation Raised By:",
        })

    # ── Register Sheet ────────────────────────────────────────────────

    def register_cell(self, field: str) -> str:
        return self.config.get("register_sheet", {}).get(field, "")

    @property
    def register_vo_header_row(self) -> int:
        return self.config.get("register_sheet", {}).get("vo_table_start_row", 12)

    @property
    def register_vo_first_data_row(self) -> int:
        return self.config.get("register_sheet", {}).get("vo_table_first_data_row", 13)

    def register_vo_col(self, field: str) -> str:
        return self.config.get("register_sheet", {}).get("vo_table_columns", {}).get(field, "A")

    # ── Helpers ───────────────────────────────────────────────────────

    def resolve_cell(self, sheet_section: str, field: str) -> str:
        """Resolve a cell reference from a section + field name."""
        section = self.config.get(sheet_section, {})
        return section.get(field, "")

    @classmethod
    def from_yaml(cls, path: Path) -> "TemplateMapping":
        """Load mapping from a YAML file."""
        with open(path, "r", encoding="utf-8") as f:
            config = yaml.safe_load(f)
        return cls(config)

    @classmethod
    def from_json(cls, path: Path) -> "TemplateMapping":
        """Load mapping from a JSON file."""
        with open(path, "r", encoding="utf-8") as f:
            config = json.load(f)
        return cls(config)

    @classmethod
    def from_db(cls, project_name: str | None = None) -> "TemplateMapping | None":
        """Load mapping from the variation_templates DB table."""
        from shared_tools.ipc_bridge import get_template_mapping as db_get
        record = db_get(project_name)
        if record and record.get("mapping"):
            return cls(record["mapping"])
        return None

    @classmethod
    def default(cls) -> "TemplateMapping":
        """Return the hardcoded Welink default mapping (fallback if no config file)."""
        return cls(_DEFAULT_MAPPING)


# =============================================================================
# Default Mapping — Welink Construction template
# =============================================================================

_DEFAULT_MAPPING: dict[str, Any] = {
    "template": {
        "blank_sheet": "VOXX",
        "register_sheet": "Register",
    },
    "vo_sheet": {
        "title_cell": "A4",
        "project_name": "B6",
        "date_issued": "G6",
        "company_name": "B7",
        "site_instruction_ref": "F7",
        "site_address": "B8",
        "job_number": "B9",
        "vo_title": "B11",
        "items_header_row": 12,
        "items_start_row": 14,
        "items_max_rows": 50,
        "items_columns": {
            "item": "A",
            "description": "B",
            "qty": "C",
            "unit": "D",
            "rate": "E",
            "cost": "F",
            "credit": "G",
        },
        "totals_labels": {
            "sub_total": "SUB TOTAL",
            "nett_cost": "NETT VARIATION COST",
            "margin": "MARGIN AND OVERHEAD COSTS",
            "excl_gst": "VARIATION COST EXCLUDING GST",
            "gst": "GST",
            "total": "TOTAL INCLUDING GST",
            "raised_by": "Variation Raised By:",
        },
        "raised_by_col": "F",
        "raised_by_value_col": "F",
        "acceptance_label_row_offset": 3,
    },
    "register_sheet": {
        "title_row": 1,
        "job_number": "B3",
        "base_contract_amount": "G3",
        "project_name": "B4",
        "project_location": "B5",
        "vo_table_start_row": 12,
        "vo_table_first_data_row": 13,
        "vo_table_columns": {
            "vo_number": "A",
            "description": "B",
            "date_issued": "C",
            "variation_value": "D",
            "pending_approval": "E",
            "not_approved": "F",
            "total_approved": "G",
            "not_proceeding": "H",
            "in_dispute": "I",
            "nod_eot": "J",
            "status": "K",
            "notes": "L",
            "action": "M",
        },
    },
}


# =============================================================================
# VariationExcelBuilder — fill template, write formulas, save
# =============================================================================

class VariationExcelBuilder:
    """Opens a master template workbook, fills the VO sheet and Register sheet,
    writes dynamic formulas, and saves the populated workbook."""

    def __init__(self, mapping: TemplateMapping, template_path: Path):
        self.mapping = mapping
        self.template_path = template_path
        self.wb: openpyxl.Workbook | None = None

    def open(self) -> None:
        """Open the template workbook."""
        self.wb = openpyxl.load_workbook(self.template_path)

    # ── VO Sheet ──────────────────────────────────────────────────────

    def create_vo_sheet(self, vo_number: int) -> str:
        """Create a fresh blank sheet named 'VO{number}'. Returns the sheet name."""
        if self.wb is None:
            raise RuntimeError("Call open() first")

        sheet_name = f"VO{vo_number}"
        if sheet_name in self.wb.sheetnames:
            del self.wb[sheet_name]

        ws = self.wb.create_sheet(title=sheet_name)
        return sheet_name

    def build_vo_sheet(self, ws, variation: dict, items: list[dict], initials: str = "AC") -> None:
        """Build an entire VO sheet from scratch — no template dependency.
        Project info, item table, summary section, raised_by, and acceptance
        are all written in pure code at calculated positions."""
        m = self.mapping
        THIN = Side(style='thin')
        MEDIUM = Side(style='medium')
        HAIR = Side(style='hair')
        FONT = Font(name='Arial', size=10)
        FONT_BOLD = Font(name='Arial', size=10, bold=True)
        FONT_TITLE = Font(name='Arial', size=12, bold=True)
        FONT_LABEL = Font(name='Arial', size=9, bold=True)
        FONT_VALUE = Font(name='Arial', size=8)
        FONT_VALUE_BOLD = Font(name='Arial', size=8, bold=True)
        FONT_SIGN = Font(name='Rage Italic', size=18)
        HEADER_FILL = PatternFill(patternType='solid', fgColor=Color(theme=8, tint=0.7999816888943144))
        FMT_MONEY = '$#,##0.00'
        ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
        ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')
        ALIGN_LEFT = Alignment(horizontal='left', vertical='center')
        ALIGN_LEFT_TOP = Alignment(horizontal='left', vertical='top')

        cols = m.config.get("vo_sheet", {}).get("items_columns", {})
        cost_col = cols.get("cost", "F")
        credit_col = cols.get("credit", "G")
        qty_col = cols.get("qty", "C")
        rate_col = cols.get("rate", "E")
        col_list = list(cols.values())  # ['A','B','C','D','E','F','G']
        # The thick separator is between Description (B-E) and Cost/Credit (F-G)
        COL_B = col_list[1]   # 'B'
        COL_E = col_list[4]   # 'E'
        COL_F = col_list[5]   # 'F'
        COL_G = col_list[6]   # 'G'

        # Border helpers
        def _hdr_border(col):
            """Header row 12 border: thick left on A/F, thick right on E/G"""
            l = MEDIUM if col in ('A', COL_F) else THIN
            r = MEDIUM if col in (COL_E, COL_G) else THIN
            return Border(left=l, right=r, top=MEDIUM, bottom=MEDIUM)

        def _subhdr_border(col):
            """Sub-header row 13 border"""
            l = MEDIUM if col in ('A', COL_F) else THIN
            r = MEDIUM if col in (COL_E, COL_G) else THIN
            return Border(left=l, right=r, top=MEDIUM, bottom=HAIR)

        def _item_border(col, top=HAIR, bottom=HAIR):
            """Item row border: thick on group edges, hair/thin inner"""
            l = MEDIUM if col in ('A', COL_F) else THIN
            r = MEDIUM if col in (COL_E, COL_G) else THIN
            return Border(left=l, right=r, top=top, bottom=bottom)

        def _summary_border(col, top=THIN, bottom=THIN):
            """Summary row border: thick left/right, variable top/bottom"""
            l = MEDIUM if col in ('A', COL_F) else THIN
            r = MEDIUM if col in (COL_E, COL_G) else THIN
            return Border(left=l, right=r, top=top, bottom=bottom)

        # ── 1. Project Info (rows 4–9) ──────────────────────────────
        ws.merge_cells('A4:G4')
        title_text = "CONTRACT VARIATION - ESTIMATE" if variation.get("is_estimate", 0) else "CONTRACT VARIATION"
        c = ws['A4']
        c.value = title_text
        c.font = FONT_TITLE
        c.alignment = ALIGN_CENTER

        ws['A6'] = "PROJECT:"; ws['A6'].font = FONT; ws['A6'].alignment = ALIGN_LEFT
        _set_cell(ws, 'B6', variation.get("project_name", "")); ws['B6'].font = FONT; ws['B6'].alignment = ALIGN_LEFT
        _set_cell(ws, 'F6', "Date:"); ws['F6'].font = FONT; ws['F6'].alignment = ALIGN_RIGHT
        date_str = variation.get("date_issued", "")
        if date_str:
            from datetime import datetime
            try:
                dt = datetime.fromisoformat(date_str.replace("Z", "+00:00"))
                _set_cell(ws, 'G6', dt); ws['G6'].number_format = "DD/MM/YYYY"
            except (ValueError, TypeError):
                _set_cell(ws, 'G6', date_str)
        ws['G6'].font = FONT; ws['G6'].alignment = ALIGN_LEFT

        ws['A7'] = "NAME:"; ws['A7'].font = FONT; ws['A7'].alignment = ALIGN_LEFT
        _set_cell(ws, 'B7', variation.get("company_name", "Welink Construction")); ws['B7'].font = FONT; ws['B7'].alignment = ALIGN_LEFT
        _set_cell(ws, 'F7', "Site Instruction / Ref:"); ws['F7'].font = FONT; ws['F7'].alignment = ALIGN_RIGHT
        _set_cell(ws, 'G7', variation.get("site_instruction_ref", "")); ws['G7'].font = FONT; ws['G7'].alignment = ALIGN_LEFT

        ws['A8'] = "SITE ADDRESS:"; ws['A8'].font = FONT; ws['A8'].alignment = ALIGN_LEFT
        _set_cell(ws, 'B8', variation.get("project_location", "")); ws['B8'].font = FONT; ws['B8'].alignment = ALIGN_LEFT

        ws['A9'] = "JOB No:"; ws['A9'].font = FONT; ws['A9'].alignment = ALIGN_LEFT
        _set_cell(ws, 'B9', variation.get("job_number", "")); ws['B9'].font = FONT; ws['B9'].alignment = ALIGN_LEFT

        # VO title
        _set_cell(ws, 'B11', f"VO{variation.get('vo_number', '')} - {variation.get('vo_title', '')}")
        ws['B11'].font = FONT_BOLD
        ws['B11'].border = Border(bottom=MEDIUM)

        # ── 2. Item Table Header (rows 12–13) ──────────────────────
        header_cells_12 = [
            ('A12', 'Item', _hdr_border('A'), ALIGN_CENTER),
            ('F12', 'Cost', _hdr_border(COL_F), ALIGN_CENTER),
        ]
        for ref, val, border, align in header_cells_12:
            c = ws[ref]; c.value = val; c.font = FONT_BOLD; c.border = border
            c.alignment = align; c.fill = HEADER_FILL
        # Description: B12:E12 merged — set borders + fill on all merged cells
        ws.merge_cells('B12:E12')
        for col in ['B', 'C', 'D', 'E']:
            c = ws[f'{col}12']
            c.border = Border(left=THIN if col != 'E' else MEDIUM,
                             right=MEDIUM if col == 'E' else THIN,
                             top=MEDIUM, bottom=MEDIUM)
            c.fill = HEADER_FILL
        c = ws['B12']; c.value = 'Description'; c.font = FONT_BOLD; c.alignment = ALIGN_CENTER
        # Credit
        c = ws['G12']; c.value = 'Credit'; c.font = FONT_BOLD
        c.border = Border(left=THIN, right=MEDIUM, top=MEDIUM, bottom=MEDIUM)
        c.alignment = ALIGN_CENTER; c.fill = HEADER_FILL

        # Row 13 sub-headers — borders only, no fill
        sub_cols = {'C': 'Qty', 'D': 'Unit', 'E': 'Rate ($)'}
        for col_letter, label in sub_cols.items():
            c = ws[f'{col_letter}13']; c.value = label; c.font = FONT_BOLD
            c.border = _subhdr_border(col_letter); c.alignment = ALIGN_CENTER
        for col_letter in ['A', 'B', COL_F, COL_G]:
            c = ws[f'{col_letter}13']; c.border = _subhdr_border(col_letter)

        # Column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 42
        for col_letter in ['C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col_letter].width = 14

        # ── 3. Item Rows (dynamic, starting at row 14) ─────────────
        ITEM_START = 14
        items_count = len(items)
        for i in range(items_count):
            row = ITEM_START + i
            is_last = (i == items_count - 1)
            for col_letter in col_list:
                c = ws[f'{col_letter}{row}']
                c.border = _item_border(col_letter,
                    top=HAIR if i > 0 else HAIR,
                    bottom=MEDIUM if is_last else HAIR)
                c.font = FONT
                c.alignment = ALIGN_CENTER
            # Item number
            ws[f'A{row}'] = i + 1
            # Description
            ws[f'B{row}'] = items[i].get("description", "")
            ws[f'B{row}'].alignment = ALIGN_LEFT
            # Qty, Unit, Rate
            _set_cell(ws, f'{qty_col}{row}', items[i].get("qty", 0))
            _set_cell(ws, f'{cols.get("unit", "D")}{row}', items[i].get("unit", "item"))
            _set_cell(ws, f'{rate_col}{row}', items[i].get("rate", 0))
            # Cost formula with $ format
            ws[f'{cost_col}{row}'] = f'={qty_col}{row}*{rate_col}{row}'
            ws[f'{cost_col}{row}'].number_format = FMT_MONEY
            ws[f'{cost_col}{row}'].alignment = ALIGN_RIGHT
            # Credit with $ format
            _set_cell(ws, f'{credit_col}{row}', items[i].get("credit", 0))
            ws[f'{credit_col}{row}'].number_format = FMT_MONEY
            ws[f'{credit_col}{row}'].alignment = ALIGN_RIGHT

        last_item_row = ITEM_START + items_count - 1

        # ── 4. Summary Section (2 blank rows below last item) ──────
        SUMMARY = last_item_row + 3  # +1 gap, +2 gap, then first summary row
        summary_rows = [
            ("SUB TOTAL",
             f"=SUM({cost_col}{ITEM_START}:{cost_col}{last_item_row})",
             f"=SUM({credit_col}{ITEM_START}:{credit_col}{last_item_row})",
             False),
            ("NETT VARIATION COST",
             f"={cost_col}{SUMMARY}-{credit_col}{SUMMARY}", "", True),
            ("MARGIN AND OVERHEAD COSTS (10%)",
             f"={cost_col}{SUMMARY + 1}*0.1", "", False),
            ("VARIATION COST EXCLUDING GST",
             f"={cost_col}{SUMMARY + 1}+{cost_col}{SUMMARY + 2}", "", False),
            ("GST",
             f"={cost_col}{SUMMARY + 3}*0.1", "", False),
            ("TOTAL INCLUDING GST",
             f"={cost_col}{SUMMARY + 3}+{cost_col}{SUMMARY + 4}", "", True),
        ]

        for i, (label, cost_formula, credit_formula, is_key) in enumerate(summary_rows):
            row = SUMMARY + i
            # Label column A — plain text, no borders
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = FONT_LABEL
            # Cost value column F — $ format, right-aligned
            ws[f'{cost_col}{row}'] = cost_formula
            ws[f'{cost_col}{row}'].font = FONT_VALUE_BOLD if is_key else FONT_VALUE
            ws[f'{cost_col}{row}'].alignment = ALIGN_RIGHT
            ws[f'{cost_col}{row}'].number_format = FMT_MONEY
            # Credit value (SUB TOTAL only)
            if credit_formula:
                ws[f'{credit_col}{row}'] = credit_formula
                ws[f'{credit_col}{row}'].font = FONT_VALUE
                ws[f'{credit_col}{row}'].alignment = ALIGN_RIGHT
                ws[f'{credit_col}{row}'].number_format = FMT_MONEY

        summary_end = SUMMARY + len(summary_rows) - 1

        # ── 5. Signature Box (2 rows below summary) ─────────────────
        # Thick-bordered F:G box: Variation Raised By → initials → Authorised By
        # ACCEPTED FOR... sits in column A on the same row as Authorised By
        box_top = summary_end + 3
        rb_row = box_top       # Variation Raised By:
        init_row = rb_row + 1  # initials
        acc_row = init_row + 2 # Authorised By: (F) + ACCEPTED FOR... (A)
        box_bottom = acc_row + 3  # 3 cells below Authorised By

        # Fill the box with borders: thick top/bottom/left/right
        for row in range(box_top, box_bottom + 1):
            for col_letter in ['F', 'G']:
                c = ws[f'{col_letter}{row}']
                l = MEDIUM if col_letter == 'F' else THIN
                r = MEDIUM if col_letter == 'G' else THIN
                t = MEDIUM if row == box_top else THIN
                b = MEDIUM if row == box_bottom else THIN
                c.border = Border(left=l, right=r, top=t, bottom=b)

        # Variation Raised By: label
        ws[f'F{rb_row}'] = "Variation Raised By:"
        ws[f'F{rb_row}'].font = FONT_LABEL; ws[f'F{rb_row}'].alignment = ALIGN_LEFT
        # Initials
        ws[f'F{init_row}'] = initials
        ws[f'F{init_row}'].font = FONT_SIGN; ws[f'F{init_row}'].alignment = ALIGN_CENTER
        # Authorised By: (F) + ACCEPTED FOR... (A) — same row
        ws[f'F{acc_row}'] = "Authorised By:"
        ws[f'F{acc_row}'].font = FONT_LABEL; ws[f'F{acc_row}'].alignment = ALIGN_LEFT
        ws[f'A{acc_row}'] = "ACCEPTED FOR AND ON BEHALF OF CLIENT:"
        ws[f'A{acc_row}'].font = FONT_LABEL

        # ── Page Setup ──────────────────────────────────────────────
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_setup.paperSize = 9  # A4
        ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
        # Print area: A1 to G{last content row}
        last_row = box_bottom
        ws.print_area = f'A1:G{last_row}'

    # ── Register Sheet ────────────────────────────────────────────────

    def fill_register_project_info(self, ws, variation: dict) -> None:
        """Fill project-level info on the Register sheet."""
        m = self.mapping
        _set_cell(ws, m.register_cell("job_number"), variation.get("job_number", ""))
        _set_cell(ws, m.register_cell("project_name"), variation.get("project_name", ""))
        _set_cell(ws, m.register_cell("project_location"), variation.get("project_location", ""))
        _set_cell(ws, m.register_cell("base_contract_amount"), variation.get("base_contract_amount", 0))

    def fill_register_vo_row(self, ws, variation: dict, calculated: dict,
                             vo_row: int | None = None) -> int:
        """Fill one VO row in the Register sheet. If vo_row is None, finds the
        row matching this VO number. Returns the row number used."""
        m = self.mapping
        cols = m.config.get("register_sheet", {}).get("vo_table_columns", {})

        vo_number = variation.get("vo_number", 1)

        # Find the row for this VO number
        if vo_row is None:
            vo_row = _find_vo_row(ws, vo_number, cols.get("vo_number", "A"),
                                   start_row=m.register_vo_first_data_row)
            if vo_row is None:
                # Use first available empty row
                vo_row = m.register_vo_first_data_row
                data_row = m.register_vo_first_data_row
                while ws[f"{cols.get('vo_number', 'A')}{data_row}"].value is not None:
                    data_row += 1
                vo_row = data_row

        _set_cell(ws, f"{cols.get('vo_number', 'A')}{vo_row}", vo_number)
        _set_cell(ws, f"{cols.get('description', 'B')}{vo_row}", variation.get("vo_title", ""))
        _set_cell(ws, f"{cols.get('date_issued', 'C')}{vo_row}", variation.get("date_issued", ""))
        _set_cell(ws, f"{cols.get('variation_value', 'D')}{vo_row}",
                   calculated.get("total_incl_gst", 0))
        _set_cell(ws, f"{cols.get('notes', 'L')}{vo_row}", variation.get("vo_type", ""))

        # Status with formatting
        status = variation.get("status", "submitted")
        status_map = {
            "submitted": "Submitted",
            "approved": "Approved",
            "approved_for_signing": "Approved for Signing",
            "not_approved": "Not Approved",
            "void": "Void",
        }
        display_status = status_map.get(status, status.replace("_", " ").title())
        status_cell = ws[f"{cols.get('status', 'K')}{vo_row}"]
        status_cell.value = display_status
        status_cell.alignment = Alignment(horizontal="left", vertical="center")

        # Color fill for status cells
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        if status in ("approved", "approved_for_signing"):
            status_cell.fill = green_fill
        elif status == "submitted":
            status_cell.fill = yellow_fill
        elif status == "not_approved":
            status_cell.fill = red_fill

        # Set pending/approved values based on status
        total = calculated.get("total_incl_gst", 0)
        if status == "submitted":
            _set_cell(ws, f"{cols.get('pending_approval', 'E')}{vo_row}", total)
            _set_cell(ws, f"{cols.get('total_approved', 'G')}{vo_row}", 0)
            _set_cell(ws, f"{cols.get('not_approved', 'F')}{vo_row}", 0)
        elif status in ("approved", "approved_for_signing"):
            stored_appr = _safe_float(variation.get("approved_value", 0))
            stored_not = _safe_float(variation.get("not_approved_value", 0))
            approved = stored_appr if stored_appr > 0 else total
            not_appr = stored_not
            _set_cell(ws, f"{cols.get('pending_approval', 'E')}{vo_row}", 0)
            _set_cell(ws, f"{cols.get('total_approved', 'G')}{vo_row}", approved)
            _set_cell(ws, f"{cols.get('not_approved', 'F')}{vo_row}", not_appr)
        elif status == "not_approved":
            stored_appr = _safe_float(variation.get("approved_value", 0))
            stored_not = _safe_float(variation.get("not_approved_value", 0))
            approved = stored_appr
            not_appr = stored_not if stored_not > 0 else total
            _set_cell(ws, f"{cols.get('pending_approval', 'E')}{vo_row}", 0)
            _set_cell(ws, f"{cols.get('total_approved', 'G')}{vo_row}", approved)
            _set_cell(ws, f"{cols.get('not_approved', 'F')}{vo_row}", not_appr)
        elif status == "void":
            _set_cell(ws, f"{cols.get('pending_approval', 'E')}{vo_row}", 0)
            _set_cell(ws, f"{cols.get('total_approved', 'G')}{vo_row}", 0)
            _set_cell(ws, f"{cols.get('not_approved', 'F')}{vo_row}", 0)

        return vo_row

    def update_register_totals(self, ws, variations: list[dict]) -> None:
        """Recalculate the Register sheet totals row for all VOs."""
        m = self.mapping
        cols = m.config.get("register_sheet", {}).get("vo_table_columns", {})
        first_row = m.register_vo_first_data_row

        # Find the totals row
        totals_row = _find_label_row(ws, "TOTALS", start_col="B", start_row=first_row, max_scan=50)
        if not totals_row:
            return

        # Sum all VO data rows
        total_value = 0
        total_pending = 0
        total_approved = 0
        total_not_approved = 0
        row = first_row
        while row < totals_row:
            vo_num = ws[f"{cols.get('vo_number', 'A')}{row}"].value
            if vo_num is not None:
                total_value += _safe_float(ws[f"{cols.get('variation_value', 'D')}{row}"].value)
                total_pending += _safe_float(ws[f"{cols.get('pending_approval', 'E')}{row}"].value)
                total_not_approved += _safe_float(ws[f"{cols.get('not_approved', 'F')}{row}"].value)
                total_approved += _safe_float(ws[f"{cols.get('total_approved', 'G')}{row}"].value)
            row += 1

        _set_cell(ws, f"{cols.get('variation_value', 'D')}{totals_row}", total_value)
        _set_cell(ws, f"{cols.get('pending_approval', 'E')}{totals_row}", total_pending)
        _set_cell(ws, f"{cols.get('not_approved', 'F')}{totals_row}", total_not_approved)
        _set_cell(ws, f"{cols.get('total_approved', 'G')}{totals_row}", total_approved)

    # ── Save ──────────────────────────────────────────────────────────

    def save(self, output_path: Path) -> None:
        """Save the modified workbook to the output path."""
        if self.wb is None:
            raise RuntimeError("Call open() first")
        output_path.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(str(output_path))

    def close(self) -> None:
        if self.wb:
            self.wb.close()
            self.wb = None


# =============================================================================
# Helpers
# =============================================================================

def _set_cell(ws, cell_ref: str, value: Any) -> None:
    """Set a cell value if cell_ref is not empty."""
    if cell_ref:
        try:
            ws[cell_ref] = value
        except Exception:
            pass


def _unmerge_range(ws, start_row: int, end_row: int, col_letters: list[str]) -> None:
    """Unmerge any merged cells that overlap the given row/column range."""
    merged_to_unmerge = []
    for merged_range in list(ws.merged_cells.ranges):
        mr_min_row = merged_range.min_row
        mr_max_row = merged_range.max_row
        if mr_max_row < start_row or mr_min_row > end_row:
            continue
        merged_to_unmerge.append(str(merged_range))
    for ref in merged_to_unmerge:
        try:
            ws.unmerge_cells(ref)
        except Exception:
            pass


def _copy_row_format(ws, source_row: int, target_row: int, col_letters: list[str]) -> None:
    """Copy cell formatting (font, alignment, fill, number_format — NOT border)
    from source_row to target_row for all given column letters.
    Borders are handled separately to ensure thin inner lines."""
    from copy import copy
    for col in col_letters:
        try:
            src_cell = ws[f"{col}{source_row}"]
            dst_cell = ws[f"{col}{target_row}"]
            if src_cell.has_style:
                dst_cell.font = copy(src_cell.font)
                dst_cell.fill = copy(src_cell.fill)
                dst_cell.number_format = src_cell.number_format
                dst_cell.alignment = copy(src_cell.alignment)
        except Exception:
            pass


def _safe_float(value: Any) -> float:
    """Convert a value to float, returning 0 on failure."""
    try:
        return float(value) if value is not None else 0.0
    except (ValueError, TypeError):
        return 0.0


def _find_label_row(ws, label: str, start_col: str = "A", start_row: int = 1,
                    max_scan: int = 50) -> int | None:
    """Scan a column for a matching label and return the row number.
    Performs case-insensitive partial match."""
    label_lower = label.lower().strip()
    for i in range(max_scan):
        row = start_row + i
        cell_val = ws[f"{start_col}{row}"].value
        if cell_val and label_lower in str(cell_val).lower().strip():
            return row
    return None


def _find_vo_row(ws, vo_number: int, col: str = "A", start_row: int = 13) -> int | None:
    """Find the row in a register sheet that contains a given VO number."""
    for i in range(30):
        row = start_row + i
        cell_val = ws[f"{col}{row}"].value
        try:
            if int(cell_val) == vo_number:
                return row
        except (ValueError, TypeError):
            if cell_val is None:
                return None  # reached empty rows
    return None


# =============================================================================
# Deterministic Cost Calculation (NO LLM — per CLAUDE.md directive)
# =============================================================================

def calculate_variation_costs(items: list[dict]) -> dict:
    """Pure Python math — never involves LLM.

    Returns a dict with:
        sub_total, credits, nett_variation_cost, margin,
        excl_gst, gst, total_incl_gst
    """
    for item in items:
        qty = _safe_float(item.get("qty", 0))
        rate = _safe_float(item.get("rate", 0))
        item["cost"] = round(qty * rate, 2)
        item["credit"] = _safe_float(item.get("credit", 0))

    sub_total = round(sum(_safe_float(i.get("cost", 0)) for i in items), 2)
    total_credits = round(sum(_safe_float(i.get("credit", 0)) for i in items), 2)
    nett = round(sub_total - total_credits, 2)
    margin = round(nett * 0.10, 2)          # 10% margin & overhead
    excl_gst = round(nett + margin, 2)
    gst = round(excl_gst * 0.10, 2)          # 10% GST
    total = round(excl_gst + gst, 2)

    return {
        "sub_total": sub_total,
        "credits": total_credits,
        "nett_variation_cost": nett,
        "margin": margin,
        "excl_gst": excl_gst,
        "gst": gst,
        "total_incl_gst": total,
    }


# =============================================================================
# Import Project from xlsx — parse existing file into project + variations
# =============================================================================

def import_project_from_xlsx(xlsx_path: str, mapping: TemplateMapping | None = None) -> dict:
    """Parse an existing variation xlsx file and extract project + all VOs + items.

    Returns:
        {
            "project": {name, job_number, location, base_contract_amount, ...},
            "variations": [{entry_id, vo_number, vo_title, ..., items: [...]}, ...]
        }
    """
    import uuid
    from datetime import datetime

    if mapping is None:
        mapping = TemplateMapping.default()

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # ── 1. Read Register sheet → project info ─────────────────────────
    project = {}
    register_name = mapping.register_sheet_name
    if register_name in wb.sheetnames:
        ws = wb[register_name]
        project["name"] = _read_cell(ws, mapping.register_cell("project_name")) or ""
        project["job_number"] = str(_read_cell(ws, mapping.register_cell("job_number")) or "")
        project["location"] = _read_cell(ws, mapping.register_cell("project_location")) or ""
        project["base_contract_amount"] = _safe_float(
            _read_cell(ws, mapping.register_cell("base_contract_amount")))
        project["company_name"] = "Welink Construction"
        project["xlsx_path"] = str(Path(xlsx_path).resolve())
        project["source_type"] = "imported"

    # ── 2. Read all VO sheets ─────────────────────────────────────────
    variations = []
    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("VO") or sheet_name == "VOXX":
            continue

        # Extract VO number from sheet name
        try:
            vo_number = int(sheet_name.replace("VO", ""))
        except ValueError:
            continue

        ws = wb[sheet_name]
        vo_title = _read_cell(ws, mapping.vo_cell("vo_title")) or f"VO{vo_number}"

        # Read date
        date_val = _read_cell(ws, mapping.vo_cell("date_issued"))
        date_issued = None
        if isinstance(date_val, datetime):
            date_issued = date_val.isoformat()
        elif date_val:
            date_issued = str(date_val)

        # Detect estimate
        title_cell = mapping.vo_cell("title_cell")
        title_text = str(_read_cell(ws, title_cell) or "")
        is_estimate = 1 if "ESTIMATE" in title_text.upper() else 0

        # Read line items
        items = []
        start_row = mapping.vo_items_start_row
        max_rows = mapping.vo_items_max_rows
        cols = mapping.config.get("vo_sheet", {}).get("items_columns", {})
        item_col = cols.get("item", "A")
        desc_col = cols.get("description", "B")
        qty_col = cols.get("qty", "C")
        unit_col = cols.get("unit", "D")
        rate_col = cols.get("rate", "E")
        cost_col = cols.get("cost", "F")
        credit_col = cols.get("credit", "G")

        for i in range(max_rows):
            row = start_row + i
            item_num = _read_cell(ws, f"{item_col}{row}")
            if item_num is None:
                continue
            try:
                item_number = int(item_num)
            except (ValueError, TypeError):
                continue

            description = _read_cell(ws, f"{desc_col}{row}") or ""
            if not description:
                continue

            qty = _safe_float(_read_cell(ws, f"{qty_col}{row}"))
            unit = _read_cell(ws, f"{unit_col}{row}") or "item"
            rate = _safe_float(_read_cell(ws, f"{rate_col}{row}"))
            cost = _safe_float(_read_cell(ws, f"{cost_col}{row}"))
            credit = _safe_float(_read_cell(ws, f"{credit_col}{row}"))

            items.append({
                "item_number": item_number,
                "description": str(description),
                "qty": qty,
                "unit": str(unit),
                "rate": rate,
                "cost": cost,
                "credit": credit,
                "sort_order": i,
            })

        # Detect status from totals
        status = "draft"
        # Check if totals are filled (non-zero) → likely submitted/approved
        totals = calculate_variation_costs(items) if items else {}
        if totals.get("total_incl_gst", 0) > 0:
            status = "submitted"

        variations.append({
            "entry_id": str(uuid.uuid4()),
            "vo_number": vo_number,
            "vo_title": str(vo_title),
            "vo_type": "Head Contract VO",
            "is_estimate": is_estimate,
            "date_issued": date_issued,
            "status": status,
            "items": items,
            "project_name": project.get("name", ""),
            "project_location": project.get("location", ""),
            "job_number": project.get("job_number", ""),
            "base_contract_amount": project.get("base_contract_amount", 0),
        })

    # ── 3. Read Internal VO Register → bank_approved / client_approved ─
    internal_register_name = "Internal VO Register"
    if internal_register_name in wb.sheetnames:
        iws = wb[internal_register_name]
        # Scan rows for VO data
        for row in range(4, 30):
            vo_num = _read_cell(iws, f"A{row}")
            if vo_num is None:
                continue
            try:
                vo_num = int(vo_num)
            except (ValueError, TypeError):
                if row > 4 and _read_cell(iws, f"B{row}") is None:
                    break  # passed the data section
                continue

            bank = _safe_float(_read_cell(iws, f"D{row}"))
            client = _safe_float(_read_cell(iws, f"E{row}"))

            # Match to variation
            for var in variations:
                if var["vo_number"] == vo_num:
                    var["bank_approved"] = bank
                    var["client_approved"] = client
                    break

    wb.close()
    return {"project": project, "variations": variations}


# =============================================================================
# Compile Project to xlsx — write all VOs + Registers to file
# =============================================================================

def compile_project_to_xlsx(project: dict, variations: list[dict],
                            output_path: Path, mapping: TemplateMapping | None = None) -> Path:
    """Compile a complete project xlsx with all VO sheets + Register + Internal VO Register.

    VOs are ordered before Register/Internal VO Register. Voided VOs are excluded
    and their sheets removed from the xlsx.
    """
    if mapping is None:
        mapping = TemplateMapping.default()

    # ── Filter: exclude voided VOs ────────────────────────────────────
    active_variations = [v for v in variations if v.get("status") != "void"]
    voided_vo_numbers = {v.get("vo_number") for v in variations if v.get("status") == "void"}

    # Sort by sort_order (drag-and-drop sequence), then vo_number
    active_variations.sort(key=lambda v: (v.get("sort_order", 0) or 0, v.get("vo_number", 0) or 0))

    # Find template
    knowledge_dir = Path(__file__).parent.parent.parent.parent / "knowledge"
    cleaned = knowledge_dir / "variation_template.xlsx"
    original = knowledge_dir / "drafted simple workflow" / "20260602 47CBR - Welink Construction Client Variations.xlsx"
    template_path = cleaned if cleaned.exists() else original

    if not template_path.exists():
        raise FileNotFoundError(f"Template not found at {template_path}")

    builder = VariationExcelBuilder(mapping, template_path)
    builder.open()

    # ── Remove old VO sheets (keep VOXX for cloning new ones) ─────────
    for name in list(builder.wb.sheetnames):
        if name.startswith("VO") and name != "VOXX":
            del builder.wb[name]

    # ── 2. Create/fill VO sheets ──
    for var in active_variations:
        # Parse VO number from title prefix (e.g., "VO3 - Description" → 3)
        vo_title = var.get("vo_title", "")
        import re
        vo_match = re.match(r'VO(\d+)', vo_title)
        vo_number = int(vo_match.group(1)) if vo_match else (var.get("vo_number") or 1)
        sheet_name = builder.create_vo_sheet(vo_number)
        ws = builder.wb[sheet_name]

        # Override with project-level fields (not per-VO denormalized copies)
        var_for_sheet = dict(var)
        var_for_sheet["project_name"] = project.get("name", var.get("project_name", ""))
        var_for_sheet["project_location"] = project.get("location", var.get("project_location", ""))
        var_for_sheet["job_number"] = project.get("job_number", var.get("job_number", ""))
        var_for_sheet["company_name"] = project.get("company_name", var.get("company_name", "Welink Construction"))

        builder.build_vo_sheet(ws, var_for_sheet, var.get("items", []),
                               initials=var.get("raised_by", "AC"))

    # ── 3. Reorder: VO sheets first (in creation order = sort_order), then Register, Internal VO Register ──
    # VO sheets were created in active_variations order (sorted by sort_order).
    # Preserve that order — do NOT re-sort by VO number.
    vo_sheets = [s for s in builder.wb.sheetnames if s.startswith("VO") and s != "VOXX"]
    trailing = ["Register", "Internal VO Register"]
    trailing = [s for s in trailing if s in builder.wb.sheetnames]
    trailing += [s for s in builder.wb.sheetnames
                 if s not in vo_sheets and s not in trailing]

    # Move all trailing sheets to the end
    for name in trailing:
        current = builder.wb.sheetnames.index(name)
        end = len(builder.wb.sheetnames) - 1
        builder.wb.move_sheet(name, offset=end - current)

    # Move VO sheets to the front in reverse order, preserving their relative sequence
    for name in reversed(vo_sheets):
        builder.wb.move_sheet(name, offset=-builder.wb.sheetnames.index(name))

    # ── 4. Write Register sheet ───────────────────────────────────────
    register_name = mapping.register_sheet_name
    if register_name in builder.wb.sheetnames:
        reg_ws = builder.wb[register_name]
        builder.fill_register_project_info(reg_ws, project)

        # Clear ALL old VO data rows first (between header and totals)
        first_data_row = mapping.register_vo_first_data_row
        # Find the TOTALS label row or clear up to a safe maximum
        totals_row = _find_label_row(reg_ws, "TOTALS", start_col="B", start_row=first_data_row, max_scan=50)
        clear_end = totals_row - 1 if totals_row else first_data_row + 30
        for row in range(first_data_row, clear_end + 1):
            for col_letter in mapping.config.get("register_sheet", {}).get("vo_table_columns", {}).values():
                reg_ws[f"{col_letter}{row}"] = None

        for var in active_variations:
            items = var.get("items", [])
            calculated = calculate_variation_costs(items) if items else {}
            builder.fill_register_vo_row(reg_ws, var, calculated)

        builder.update_register_totals(reg_ws, active_variations)

    # ── 5. Write Internal VO Register sheet ───────────────────────────
    internal_name = "Internal VO Register"
    if internal_name in builder.wb.sheetnames:
        iws = builder.wb[internal_name]
        # Clear all old data rows
        _clear_internal_register(iws)
        _fill_internal_register(iws, active_variations)

    # ── 6. Copy images from VOXX to each VO sheet ──────────────────────
    if "VOXX" in builder.wb.sheetnames:
        # Cache image data before copying (data is consumed on first read)
        cached_images = _extract_images(builder.wb["VOXX"])
        for sheet_name in vo_sheets:
            if sheet_name in builder.wb.sheetnames:
                for img_data in cached_images:
                    _add_image_to_sheet(builder.wb[sheet_name], img_data)

    # ── 7. Set all fonts to Arial ─────────────────────────────────────
    _set_all_fonts_arial(builder.wb)

    # ── 8. Cleanup: remove VOXX template blank from output ─────────────
    if "VOXX" in builder.wb.sheetnames:
        del builder.wb["VOXX"]

    # ── 9. Save ───────────────────────────────────────────────────────
    output_path.parent.mkdir(parents=True, exist_ok=True)
    builder.save(output_path)
    builder.close()

    return output_path


def _fill_internal_register(ws, variations: list[dict]) -> None:
    """Fill the Internal VO Register sheet — mirrors Register's status-based logic.
    Columns: A=Seq#, B=Description, C=Variation Value, D=Approved, E=Pending, F=Notes"""
    data_start_row = 4
    row = data_start_row

    seq = 1
    for var in sorted(variations, key=lambda v: v.get("vo_number", 0) or 0):
        items = var.get("items", [])
        totals = calculate_variation_costs(items) if items else {}
        total = totals.get("total_incl_gst", 0)
        status = var.get("status", "submitted")

        bank = 0
        client = 0
        pending = 0

        if status == "submitted":
            pending = total
        elif status in ("approved", "approved_for_signing"):
            stored = _safe_float(var.get("approved_value", 0))
            approved_amt = stored if stored > 0 else total
            if var.get("approval_type") == "bank":
                bank = approved_amt
            else:
                client = approved_amt
        elif status == "not_approved":
            stored = _safe_float(var.get("not_approved_value", 0))
            pending = stored if stored > 0 else total

        ws[f"A{row}"] = seq
        ws[f"B{row}"] = var.get("vo_title", "")
        ws[f"C{row}"] = total
        ws[f"C{row}"].number_format = '#,##0.00'
        ws[f"D{row}"] = bank
        ws[f"D{row}"].number_format = '#,##0.00'
        ws[f"E{row}"] = client
        ws[f"E{row}"].number_format = '#,##0.00'
        ws[f"F{row}"] = pending
        ws[f"F{row}"].number_format = '#,##0.00'
        row += 1
        seq += 1

    # Totals row
    if row > data_start_row:
        last_data_row = row - 1
        ws[f"B{row}"] = "TOTALS"
        for col in ["C", "D", "E", "F"]:
            ws[f"{col}{row}"] = f"=SUM({col}{data_start_row}:{col}{last_data_row})"


def _extract_images(ws) -> list[dict]:
    """Extract image data and anchor info from a worksheet. Caches data so it
    can be reused across multiple destination sheets."""
    if not hasattr(ws, '_images') or not ws._images:
        return []
    results = []
    for img in ws._images:
        try:
            data = None
            if hasattr(img, '_data'):
                data = img._data()
            elif hasattr(img, 'ref') and img.ref:
                data = img.ref
            if data:
                anchor_info = None
                if hasattr(img, 'anchor'):
                    anchor_info = img.anchor
                results.append({
                    'data': data,
                    'width': img.width,
                    'height': img.height,
                    'anchor': anchor_info,
                })
        except Exception:
            pass
    return results


def _add_image_to_sheet(ws, img_info: dict) -> None:
    """Add a cached image to a worksheet."""
    from openpyxl.drawing.image import Image
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor
    import io
    try:
        new_img = Image(io.BytesIO(img_info['data']))
        new_img.width = img_info['width']
        new_img.height = img_info['height']
        if img_info['anchor'] and isinstance(img_info['anchor'], OneCellAnchor):
            new_anchor = OneCellAnchor()
            new_anchor._from = img_info['anchor']._from
            new_anchor.ext = img_info['anchor'].ext
            new_img.anchor = new_anchor
        ws.add_image(new_img)
    except Exception:
        pass


def _set_all_fonts_arial(wb) -> None:
    """Set all cell fonts to Arial across all sheets.
    Preserves signature-style fonts (Rage Italic, Edwardian Script) for the
    'AC' initials and other handwritten elements."""
    from openpyxl.styles import Font

    # Fonts to preserve as-is (signature/handwriting style)
    SIGNATURE_FONTS = {'Rage Italic', 'Edwardian Script ITC'}

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                fn = cell.font.name if cell.font else None
                # Skip signature cells — keep their original handwriting font
                if fn in SIGNATURE_FONTS:
                    continue
                # Update cells with content or non-Arial font
                if cell.value is not None or (fn and fn != 'Arial'):
                    old = cell.font
                    cell.font = Font(
                        name='Arial',
                        size=(old.size if old and old.size else 10),
                        bold=(old.bold if old and old.bold else False),
                        italic=(old.italic if old and old.italic else False),
                        color=(old.color if old else None),
                        underline=(old.underline if old else None),
                    )


def _clear_internal_register(ws) -> None:
    """Clear all data rows in the Internal VO Register sheet, keeping headers (rows 1-3)."""
    # Clear everything from row 4 to row 30 (covers all possible old data)
    for row in range(4, 31):
        for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]:
            ws[f"{col}{row}"] = None


def _read_cell(ws, cell_ref: str) -> Any:
    """Read a cell value, returning None if cell_ref is empty or invalid."""
    if not cell_ref:
        return None
    try:
        return ws[cell_ref].value
    except Exception:
        return None
