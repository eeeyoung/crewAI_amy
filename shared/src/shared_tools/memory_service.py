"""MemoryService — RAG ingestion & retrieval engine for lilAmy.

Follows the established lilAmy service pattern (QObject + signals).
Scans a configurable data-root directory tree, extracts text from
documents, chunks them, embeds via ChromaDB, and provides semantic
search with optional project-level filtering.

Usage:
    service = MemoryService(data_root="/path/to/LILAMY_DATA_DIR")
    service.ingest_all()                # scan & index everything
    results = service.search("concrete pour schedule", project="ARCO")
"""

import gc
import os
import hashlib
import threading
from pathlib import Path

from PyQt6.QtCore import QObject, pyqtSignal

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

CHUNK_SIZE = 500       # characters per chunk
CHUNK_OVERLAP = 100    # overlap between consecutive chunks
COLLECTION_NAME = "lilamy_docs"
EMBEDDING_BATCH_SIZE = 200  # flush to ChromaDB every N chunks (reduces I/O)
EMBEDDING_GC_EVERY = 1000   # run gc.collect() every N chunks (frees PyTorch memory)

# ---------------------------------------------------------------------------
# Persistent embedding function (singleton) — strict ONNX configuration
# ---------------------------------------------------------------------------
# ChromaDB's DefaultEmbeddingFunction creates a new ONNXMiniLM_L6_V2 on every
# call, which in turn creates a fresh ONNX Runtime InferenceSession with
# enable_cpu_mem_arena=True (default).  Over 90+ upsert batches the unreleased
# arena memory can balloon into multiple GB and OOM-crash a laptop.
#
# Even the ONNXMiniLM_L6_V2.model @cached_property creates its own
# SessionOptions() internally and ignores any _user_config — there is no hook
# for custom options.  So we build the InferenceSession ourselves with strict
# options and pre-populate instance.__dict__["model"], which short-circuits
# the @cached_property so the parent's SessionOptions code never runs.
#
# We also limit intra/inter thread counts to leave CPU cores free for
# simultaneous CrewAI agents (no deadlocks during parallel agent runs).

_PERSISTENT_EF = None


def _get_persistent_embedding_function():
    """Return a module-level ONNXMiniLM_L6_V2 singleton with:
    - ``enable_cpu_mem_arena=False``  (prevents RAM bloat / OOM crash)
    - ``intra_op_num_threads=2``      (leaves cores for CrewAI agents)
    - ``inter_op_num_threads=2``
    """
    global _PERSISTENT_EF
    if _PERSISTENT_EF is not None:
        return _PERSISTENT_EF

    import onnxruntime as ort
    from chromadb.utils.embedding_functions import ONNXMiniLM_L6_V2

    instance = ONNXMiniLM_L6_V2(preferred_providers=["CPUExecutionProvider"])

    # Ensure the ONNX model files are downloaded (normally happens lazily
    # inside __call__, but we need them on disk before creating the session).
    instance._download_model_if_not_exists()

    # ── Build optimized SessionOptions ──────────────────────────────
    so = ort.SessionOptions()
    so.log_severity_level = 3
    so.graph_optimization_level = ort.GraphOptimizationLevel.ORT_ENABLE_ALL
    so.enable_cpu_mem_arena = False   # ← THE KEY FIX: return scratch memory to OS
    so.intra_op_num_threads = 2       # ← leave CPU cores for CrewAI agents
    so.inter_op_num_threads = 2

    # ── Pre-populate the @cached_property so ChromaDB's own model()
    #     (which would create a fresh SessionOptions) never runs ─────
    model_path = os.path.join(
        instance.DOWNLOAD_PATH,
        instance.EXTRACTED_FOLDER_NAME,
        "model.onnx",
    )
    session = ort.InferenceSession(
        model_path,
        providers=instance._preferred_providers,
        sess_options=so,
    )
    instance.__dict__["model"] = session

    _PERSISTENT_EF = instance
    return _PERSISTENT_EF


# ---------------------------------------------------------------------------
# MemoryService
# ---------------------------------------------------------------------------

class MemoryService(QObject):
    """Document ingestion, embedding, and semantic-search engine."""

    # -- signals --
    ingest_progress = pyqtSignal(str)    # human-readable progress message
    ingest_done = pyqtSignal(int)        # total files processed
    search_ready = pyqtSignal()          # emitted after first successful ingest

    def __init__(self, data_root: str | None = None, parent: QObject | None = None):
        super().__init__(parent)
        self._data_root = Path(data_root or os.environ.get("LILAMY_DATA_DIR", ""))
        self._lock = threading.Lock()
        self._collection = None
        self._initialised = False
        # Batch buffer — accumulate chunks and flush in one ChromaDB call
        self._buffer: dict[str, list] = {"ids": [], "documents": [], "metadatas": []}
        # In-memory hash tracker to skip already-ingested files (avoids DB query per file)
        self._seen_hashes: dict[str, str] = {}  # file_name → md5_hash

    # ------------------------------------------------------------------
    # Initialisation
    # ------------------------------------------------------------------

    @property
    def data_root(self) -> Path:
        return self._data_root

    @property
    def is_ready(self) -> bool:
        return self._initialised

    def _ensure_collection(self):
        """Lazy-init ChromaDB collection.

        Uses the persistent ONNX EF singleton (one InferenceSession total) and
        monkey-patches ONNX Runtime to disable ``enable_cpu_mem_arena`` (see
        ``_patch_onnx_arena``).  Without both fixes the arena balloons across
        batches and OOM-crashes laptops.

        If the on-disk collection was created with the old leaky
        ``DefaultEmbeddingFunction``, we nuke it and rebuild with the
        persistent EF."""
        if self._collection is not None:
            return
        import shutil
        import chromadb
        from chromadb.config import Settings

        db_dir = self._data_root / ".chromadb"
        db_dir.mkdir(parents=True, exist_ok=True)

        client = chromadb.PersistentClient(
            path=str(db_dir),
            settings=Settings(anonymized_telemetry=False),
        )
        ef = _get_persistent_embedding_function()
        try:
            self._collection = client.get_or_create_collection(
                name=COLLECTION_NAME,
                metadata={"hnsw:space": "cosine"},
                embedding_function=ef,
            )
        except ValueError:
            # Stale collection with old DefaultEmbeddingFunction — nuke & rebuild.
            # We must delete the directory on disk because ChromaDB's API-level
            # delete_collection can fail when EF classes don't match.
            self.ingest_progress.emit(
                "Migration: replacing old collection (DefaultEmbeddingFunction → ONNXMiniLM_L6_V2)"
            )
            try:
                client.delete_collection(COLLECTION_NAME)
            except Exception:
                pass
            # Belt-and-suspenders: if the API delete didn't work, remove the
            # sqlite database files directly so create_collection succeeds.
            if db_dir.exists():
                shutil.rmtree(db_dir, ignore_errors=True)
                db_dir.mkdir(parents=True, exist_ok=True)
                # Re-create the client pointing at the fresh directory
                client = chromadb.PersistentClient(
                    path=str(db_dir),
                    settings=Settings(anonymized_telemetry=False),
                )
            self._collection = client.create_collection(
                name=COLLECTION_NAME,
                metadata={"hnsw:space": "cosine"},
                embedding_function=ef,
            )

    def _flush_buffer(self):
        """Flush accumulated chunks to ChromaDB in one batch upsert."""
        if not self._buffer["ids"]:
            return
        self._ensure_collection()
        batch_size = len(self._buffer["ids"])
        self._collection.upsert(
            ids=self._buffer["ids"],
            documents=self._buffer["documents"],
            metadatas=self._buffer["metadatas"],
        )
        self._buffer = {"ids": [], "documents": [], "metadatas": []}
        return batch_size

    def _gc_check(self, chunk_count: int):
        """Periodically free memory to prevent unbounded growth."""
        if chunk_count > 0 and chunk_count % EMBEDDING_GC_EVERY == 0:
            gc.collect()
            try:
                import torch
                if torch.cuda.is_available():
                    torch.cuda.empty_cache()
            except ImportError:
                pass

    # ------------------------------------------------------------------
    # Ingestion
    # ------------------------------------------------------------------

    def ingest_directory(self, project: str | None = None) -> int:
        """Walk *data_root* and index all supported files.

        If *project* is given, only that subdirectory is processed.
        Returns the number of files processed."""
        if not self._data_root.exists():
            raise FileNotFoundError(f"Data root not found: {self._data_root}")

        self._ensure_collection()
        root = self._data_root / project if project else self._data_root
        if not root.exists():
            raise FileNotFoundError(f"Project directory not found: {root}")

        self.ingest_progress.emit(f"Scanning {root} ...")

        files = list(root.rglob("*"))
        files = [f for f in files if f.is_file() and _is_supported(f.suffix)]
        total_files = len(files)
        self.ingest_progress.emit(f"Found {total_files} supported files"
                                  f"  (batch={EMBEDDING_BATCH_SIZE},"
                                  f"  chunk={CHUNK_SIZE}chars,"
                                  f"  overlap={CHUNK_OVERLAP}chars)")

        processed = 0
        total_chunks = 0
        for idx, fp in enumerate(files, 1):
            try:
                n_chunks = self._ingest_file(fp, root)
                if n_chunks > 0:
                    processed += 1
                    total_chunks += n_chunks

                # Flush when buffer reaches batch size
                if len(self._buffer["ids"]) >= EMBEDDING_BATCH_SIZE:
                    flushed = self._flush_buffer()
                    total_chunks += 0  # already counted; just log
                    self._gc_check(total_chunks)

                # Progress every 100 files
                if idx % 100 == 0:
                    pct = idx / total_files * 100
                    # Estimate: ~10K chunks = ~30MB of vectors
                    est_mb = total_chunks * 384 * 4 / (1024 * 1024)
                    self.ingest_progress.emit(
                        f"  ... {idx}/{total_files} files ({pct:.0f}%), "
                        f"{total_chunks} chunks, ~{est_mb:.1f}MB vectors"
                    )

            except Exception as e:
                self.ingest_progress.emit(f"  ⚠️ {fp.name}: {e}")

        # Final flush
        self._flush_buffer()
        self._gc_check(total_chunks)

        self._initialised = True
        self.ingest_done.emit(processed)
        if processed > 0:
            self.search_ready.emit()
        self.ingest_progress.emit(
            f"Done — {processed}/{total_files} file(s) indexed, "
            f"{total_chunks} chunks total"
        )
        return processed

    def ingest_all(self) -> int:
        """Ingest everything under the data root (all projects)."""
        return self.ingest_directory(project=None)

    def ingest_email(self, email_dict: dict):
        """Index a single email dict (from MailService pipeline).
        The email is stored as a text chunk with metadata."""
        self._ensure_collection()

        subject = email_dict.get("subject", "")
        body = email_dict.get("body", "")
        sender = email_dict.get("sender", "")
        received = email_dict.get("received_time", "")
        content = f"Subject: {subject}\nFrom: {sender}\nDate: {received}\n\n{body}"

        project = _guess_project(subject, body)
        entry_id = email_dict.get("entry_id", "")
        chunks = _chunk_text(content)

        for i, chunk in enumerate(chunks):
            doc_id = f"email-{entry_id}-{i}"
            self._buffer["ids"].append(doc_id)
            self._buffer["documents"].append(chunk)
            self._buffer["metadatas"].append({
                "project": project,
                "file_name": f"{subject[:80]}.eml",
                "doc_type": "email",
                "chunk_index": i,
                "source": "amail",
            })

        # Flush if buffer is full
        if len(self._buffer["ids"]) >= EMBEDDING_BATCH_SIZE:
            self._flush_buffer()

    # ------------------------------------------------------------------
    # Retrieval
    # ------------------------------------------------------------------

    def search(self, query: str, project: str | None = None,
               top_k: int = 5, content_only: bool = False) -> list[dict]:
        """Semantic search across all indexed documents.

        Args:
            query: Natural-language search query.
            project: Optional project name filter.
            top_k: Number of results to return.
            content_only: If True, exclude metadata-only chunks (e.g. image
                PDFs, DWGs) and return only content-extracted text.

        Returns a list of dicts with keys: *text*, *project*, *file_name*,
        *file_path*, *file_id*, *doc_type*, *chunk_index*, *md5_hash*,
        *content_extracted*, *score* (lower = better for cosine distance).
        """
        self._ensure_collection()
        if self._collection.count() == 0:
            return []

        where = {}
        if project:
            where["project"] = project
        if content_only:
            where["content_extracted"] = True

        if not where:
            where = None

        # Fetch extra candidates so we can filter/demote metadata-only
        fetch_k = top_k * 3 if not content_only else top_k
        results = self._collection.query(
            query_texts=[query],
            n_results=fetch_k,
            where=where,
            include=["documents", "metadatas", "distances"],
        )

        hits = []
        ids = results.get("ids", [[]])[0]
        docs = results.get("documents", [[]])[0]
        metas = results.get("metadatas", [[]])[0]
        dists = results.get("distances", [[]])[0]

        for i in range(len(ids)):
            meta = metas[i] if i < len(metas) else {}
            hits.append({
                "text": docs[i] if i < len(docs) else "",
                "project": meta.get("project", ""),
                "file_name": meta.get("file_name", ""),
                "file_path": meta.get("file_path", ""),
                "file_id": meta.get("file_id", -1),
                "doc_type": meta.get("doc_type", ""),
                "chunk_index": meta.get("chunk_index", 0),
                "md5_hash": meta.get("md5_hash", ""),
                "content_extracted": meta.get("content_extracted", True),
                "score": dists[i] if i < len(dists) else 1.0,
            })

        # De-prioritise metadata-only chunks: content hits first, then
        # metadata-only — both groups remain sorted by relevance.
        if not content_only:
            content_hits = [h for h in hits if h["content_extracted"]]
            meta_hits = [h for h in hits if not h["content_extracted"]]
            hits = content_hits + meta_hits

        return hits[:top_k]

    def search_text(self, query: str, project: str | None = None,
                    top_k: int = 5) -> str:
        """Convenience: return search results formatted as a string
        (suitable for injecting into LLM prompts)."""
        hits = self.search(query, project=project, top_k=top_k)
        if not hits:
            return "No relevant documents found."

        lines = []
        for h in hits:
            lines.append(
                f"--- [{h['project']}] {h['file_name']} "
                f"({h['doc_type']}, score={h['score']:.3f}) ---\n"
                f"{h['text']}\n"
            )
        return "\n".join(lines)

    # ------------------------------------------------------------------
    # Management
    # ------------------------------------------------------------------

    def reset(self):
        """Delete all embeddings and start fresh."""
        self._ensure_collection()
        import chromadb
        try:
            client = self._collection._client
            client.delete_collection(COLLECTION_NAME)
        except Exception:
            pass
        self._collection = None
        self._initialised = False
        self._buffer = {"ids": [], "documents": [], "metadatas": []}
        self._seen_hashes.clear()
        self.ingest_progress.emit("Memory reset — all embeddings deleted")

    def stats(self) -> dict:
        """Return collection statistics."""
        self._ensure_collection()
        return {
            "data_root": str(self._data_root),
            "chunk_count": self._collection.count() if self._collection else 0,
            "collection_name": COLLECTION_NAME,
        }

    # ------------------------------------------------------------------
    # File-level ingestion (orchestrated by FileRegistry + ingest.py)
    # ------------------------------------------------------------------

    def ingest_file_with_id(
        self,
        file_path: Path,
        root: Path,
        file_id: int,
        project: str,
        file_hash: str,
    ) -> int:
        """Extract, chunk, and buffer a single file for embedding.

        Called by the ingestion orchestrator (``ingest.py``) after the file
        has been registered in ``FileRegistry``.  Stores *file_id* in every
        chunk's metadata so ChromaDB vectors link back to the SQLite registry
        (the "foreign key" of the hybrid relational-vector architecture).

        If text extraction returns empty (e.g. scanned/image PDFs, DWG files),
        a single **metadata-only chunk** is created from the filename, folder
        path, and project — so the file is still findable by name and location
        even though its content couldn't be extracted.

        Args:
            file_path: Absolute path to the file on disk.
            root: Data root (for computing relative path).
            file_id: Primary key from ``FileRegistry``.
            project: Project name (already detected by the orchestrator).
            file_hash: Pre-computed MD5 hash of the extracted text.

        Returns:
            Number of chunks buffered (1 for metadata-only, 0 for skipped).
        """
        text = _extract_text(file_path)
        rel_path = str(file_path.relative_to(root))
        file_name = file_path.name
        folder = str(file_path.parent.relative_to(root))

        # In-memory dedup within this session
        if self._seen_hashes.get(file_path.name) == file_hash:
            return 0
        self._seen_hashes[file_path.name] = file_hash

        if not text or not text.strip():
            # ── Metadata-only chunk for unextractable files ──────────
            # Construction projects have many image-based PDFs (scanned
            # drawings), DWGs, MSGs, etc.  We still index them by name,
            # path, and project so "find structural drawings" works.
            meta_text = (
                f"File: {file_name}\n"
                f"Project: {project}\n"
                f"Folder: {folder}\n"
                f"Type: {_doc_type(file_path.suffix)}"
            )
            doc_id = f"f{file_id}-meta-{file_hash[:8]}"
            self._buffer["ids"].append(doc_id)
            self._buffer["documents"].append(meta_text)
            self._buffer["metadatas"].append({
                "file_id": file_id,
                "project": project,
                "file_name": file_name,
                "file_path": rel_path,
                "doc_type": _doc_type(file_path.suffix),
                "chunk_index": 0,
                "md5_hash": file_hash,
                "content_extracted": False,   # signals "metadata only"
            })
            return 1

        chunks = _chunk_text(text)

        for i, chunk in enumerate(chunks):
            doc_id = f"f{file_id}-{file_hash[:8]}-{i}"
            self._buffer["ids"].append(doc_id)
            self._buffer["documents"].append(chunk)
            self._buffer["metadatas"].append({
                "file_id": file_id,           # ← FK → file_registry.file_id
                "project": project,
                "file_name": file_name,
                "file_path": rel_path,
                "doc_type": _doc_type(file_path.suffix),
                "chunk_index": i,
                "md5_hash": file_hash,
                "content_extracted": True,
            })

        return len(chunks)

    # ------------------------------------------------------------------
    # Internal
    # ------------------------------------------------------------------

    def _ingest_file(self, file_path: Path, root: Path) -> int:
        """Legacy flat-directory ingestion.  Prefer ``ingest_file_with_id``
        when a ``FileRegistry`` is available."""
        text = _extract_text(file_path)
        if not text or not text.strip():
            return 0

        file_hash = _md5(text)

        if self._seen_hashes.get(file_path.name) == file_hash:
            return 0
        self._seen_hashes[file_path.name] = file_hash

        project = _guess_project_from_path(file_path, root)
        if project is None:
            project = _guess_project(file_path.stem, text[:500])
        chunks = _chunk_text(text)

        for i, chunk in enumerate(chunks):
            doc_id = f"{file_path.stem}-{file_hash[:8]}-{i}"
            self._buffer["ids"].append(doc_id)
            self._buffer["documents"].append(chunk)
            self._buffer["metadatas"].append({
                "file_id": -1,                # no registry in legacy mode
                "project": project,
                "file_name": file_path.name,
                "file_path": str(file_path.relative_to(root)),
                "doc_type": _doc_type(file_path.suffix),
                "chunk_index": i,
                "md5_hash": file_hash,
            })

        return len(chunks)


# =========================================================================
# Helpers
# =========================================================================

SUPPORTED_EXTENSIONS = {
    ".txt", ".md", ".eml", ".json", ".csv", ".html",
    ".pdf", ".docx", ".doc", ".xlsx", ".xls", ".xlsm",
    ".msg", ".dwg", ".tif", ".png", ".skp", ".log", ".zip",
}

def _is_supported(suffix: str) -> bool:
    s = suffix.lower()
    if s in SUPPORTED_EXTENSIONS:
        return True
    # Optional: try importing optional extractors
    if s == ".pdf":
        try:
            import PyPDF2  # noqa: F401
            return True
        except ImportError:
            return False
    if s == ".docx":
        try:
            import docx  # noqa: F401
            return True
        except ImportError:
            return False
    if s in (".xlsx", ".xls"):
        try:
            import openpyxl  # noqa: F401
            return True
        except ImportError:
            return False
    return False


def _extract_text(file_path: Path) -> str | None:
    """Extract plain text from a file. Returns None on failure.

    For binary/image formats (DWG, TIF, PNG, SKP, ZIP) returns empty string
    — the caller creates a metadata-only chunk."""
    suffix = file_path.suffix.lower()

    try:
        if suffix == ".pdf":
            return _extract_pdf(file_path)
        elif suffix == ".docx":
            return _extract_docx(file_path)
        elif suffix == ".doc":
            # Old Word format — python-docx may work on some .doc files
            try:
                text = _extract_docx(file_path)
                if text and not _is_binary_garbage(text):
                    return text
            except Exception:
                pass
            # Try reading as text (sometimes works for simple .doc files)
            try:
                with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                    content = f.read()
                    if content.strip() and not _is_binary_garbage(content):
                        return content
            except Exception:
                pass
            return ""  # metadata-only — binary format
        elif suffix in (".xlsx", ".xls", ".xlsm"):
            return _extract_xlsx(file_path)
        elif suffix == ".msg":
            return _extract_msg(file_path)
        elif suffix in (".dwg", ".tif", ".png", ".skp", ".zip"):
            return ""  # binary — metadata-only chunk
        else:
            # txt, md, eml, json, csv, html, log
            with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                return f.read()
    except Exception:
        return None


def _extract_pdf(file_path: Path) -> str:
    import PyPDF2
    reader = PyPDF2.PdfReader(str(file_path))
    return "\n".join(
        page.extract_text() or "" for page in reader.pages
    )


def _extract_docx(file_path: Path) -> str:
    import docx
    doc = docx.Document(str(file_path))
    return "\n".join(p.text for p in doc.paragraphs)


def _extract_xlsx(file_path: Path) -> str:
    import openpyxl
    wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
    lines = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"[Sheet: {sheet_name}]")
        for row in ws.iter_rows(values_only=True):
            line = "\t".join(str(c) if c is not None else "" for c in row)
            if line.strip():
                lines.append(line)
    wb.close()
    return "\n".join(lines)


def _extract_msg(file_path: Path) -> str:
    """Extract text from an Outlook .msg file.  Uses ``extract_msg`` if
    available; returns empty string (metadata-only) otherwise."""
    try:
        import extract_msg
    except ImportError:
        return ""  # metadata-only — package not installed
    try:
        msg = extract_msg.Message(str(file_path))
        body = msg.body or ""
        subject = msg.subject or ""
        sender = msg.sender or ""
        date = str(msg.date) if msg.date else ""
        msg.close()
        return f"Subject: {subject}\nFrom: {sender}\nDate: {date}\n\n{body}"
    except Exception:
        return ""


def _is_binary_garbage(text: str, threshold: float = 0.3) -> bool:
    """Return True if *text* looks like binary data misread as text.

    Heuristic: if more than *threshold* fraction of characters are
    non-printable (outside the ASCII printable range + common whitespace),
    treat it as binary garbage that should be discarded."""
    if not text:
        return True
    printable = sum(
        1 for c in text
        if c in " \t\n\r" or (32 <= ord(c) <= 126) or ord(c) > 127
    )
    # Allow Unicode characters (ord > 127) — many real docs have them
    # Count truly binary chars: control chars except tab, newline, carriage return
    binary_chars = sum(
        1 for c in text
        if ord(c) < 32 and c not in "\t\n\r"
    )
    return (binary_chars / len(text)) > threshold


def _chunk_text(text: str) -> list[str]:
    """Split text into overlapping chunks, preserving paragraph boundaries."""
    paragraphs = text.split("\n")
    chunks = []
    current = ""
    for para in paragraphs:
        para = para.strip()
        if not para:
            if current:
                current += "\n"
            continue
        if len(current) + len(para) < CHUNK_SIZE:
            current += para + "\n"
        else:
            if current.strip():
                chunks.append(current.strip())
            # Overlap: carry last CHUNK_OVERLAP chars forward
            overlap = current[-CHUNK_OVERLAP:] if len(current) > CHUNK_OVERLAP else ""
            current = overlap + para + "\n"
    if current.strip():
        chunks.append(current.strip())
    return chunks if chunks else [text[:CHUNK_SIZE]]


def _md5(text: str) -> str:
    return hashlib.md5(text.encode("utf-8")).hexdigest()


def _doc_type(suffix: str) -> str:
    return {
        ".pdf": "pdf", ".docx": "docx", ".doc": "doc",
        ".xlsx": "spreadsheet", ".xls": "spreadsheet", ".xlsm": "spreadsheet",
        ".eml": "email", ".msg": "email",
        ".txt": "text", ".md": "markdown", ".json": "data", ".csv": "spreadsheet",
        ".html": "html", ".log": "log",
        ".dwg": "drawing", ".tif": "image", ".png": "image",
        ".skp": "model", ".zip": "archive",
    }.get(suffix.lower(), "other")


def _guess_project(subject: str, body: str) -> str:
    """Heuristic: try to identify project from email content."""
    text = f"{subject} {body[:500]}".lower()
    # Common project name patterns in construction emails.
    # Order matters: check more-specific patterns first (e.g. "econolodge"
    # before "elodge" which is an abbreviation).
    keywords = {
        "ARCO": [
            "arco", "2224 hood", "hood street", "mindarie", "boston quay",
            "subiaco",  # ARCO project location
        ],
        "Econolodge": [
            "econolodge", "econology", "elodge", "canning beach",
            "cbr", "47 canning", "12 new motel",
        ],
        "Odin Road": ["odin road", "93 odin"],
        "Kearns Crs": ["kearns crs", "kearns crescent", "2-4 kearns"],
        "Great Eastern Hwy": ["great eastern highway", "great eastern hwy", "85 great eastern"],
        "Willcock Street": ["willcock street", "6 willcock"],
        "Laviche": ["laviche"],
        "Welink": ["welink", "cheops"],  # corporate, not a project per se
    }
    for project, terms in keywords.items():
        for term in terms:
            if term in text:
                return project
    return "General"


def _guess_project_from_path(file_path: Path, root: Path) -> str | None:
    """Use the first subdirectory under root as the project name.

    Returns None when files are stored flat (no project subdirectory),
    signalling the caller to fall back to content-based heuristics."""
    try:
        rel = file_path.relative_to(root)
        parts = rel.parts
        if len(parts) >= 2:
            return parts[0]  # first subdirectory IS the project
        return None  # flat file — caller should use _guess_project()
    except ValueError:
        return None
