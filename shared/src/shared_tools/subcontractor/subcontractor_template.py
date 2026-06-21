"""Subcontractor Template — PO and Subcontract document generators.

Generates structured Excel documents from commitment data:
  - Purchase Order (PO) document — matches real Welink PO format
  - Subcontract agreement document

Uses openpyxl for Excel generation. The PO document follows the exact
Welink Construction template verified against real POs (7 samples).
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


# =============================================================================
# Constants
# =============================================================================

# Welink company details (same on every PO)
WELINK_NAME = "Welink Construction Pty Ltd"
WELINK_ADDRESS = "Suite 32 269 Vincent St"
WELINK_SUBURB = "Leederville WA 6007"
WELINK_ABN = "34 603 221 487"
WELINK_EMAIL = "accounts@welink.com.au"
WELINK_PHONE = "(08) 9204 3388"

# Logo path (extracted once, reused)
LOGO_PATH = Path(__file__).parent.parent.parent.parent.parent / "knowledge" / "welink_logo.jpeg"

# PO Terms & Conditions — IDENTICAL on every real Welink PO (10 clauses)
PO_TERMS = [
    ("1.", "TRADE WORKS",
     "(a) The trade contractor or supplier must carry out and complete the trade works: "
     "(i) to the reasonable satisfaction of the principal contractor; (ii) in accordance "
     "with the plans, the specifications and the law; and (iii) at the reasonable times "
     "directed by the principal contractor.\n"
     "(b) If the trade contractor or supplier discovers any inconsistency, ambiguity or "
     "discrepancy in or between the plans and the specifications, the trade contractor "
     "must immediately seek the principal contractor's direction as to the interpretation "
     "to be followed.\n"
     "(c) The trade contractor or supplier must supply everything necessary to carry out "
     "the trade works.\n"
     "(d) The trade contractor or supplier may employ or engage others to carry out some "
     "or all of the trade works. Use of sub-contractors does not relieve the trade "
     "contractor or supplier from liability under this trade contract."),
    ("2.", "INSURANCE",
     "(a) The trade contractor or supplier must take out prior to commencing, and maintain "
     "until completion of the trade works, the following: (i) workers compensation or any "
     "like insurance as required by law; (ii) public liability insurance to an amount not "
     "less than $5,000,000; and (iii) except as set out below, personal accident and "
     "disability insurance providing cover at least equivalent to that provided to an "
     "employee under insurance referred to in sub-clause 4(a)(i).\n"
     "(b) The trade contractor or supplier must, when asked by the principal contractor, "
     "produce evidence of the existence and currency of any insurances."),
    ("3.", "INDEMNITY",
     "The trade contractor or supplier indemnifies the principal contractor against: "
     "(a) Loss or damage to property (including the trade works); "
     "(b) Claims in respect of personal injury or death, arising out of, connected to or "
     "as a consequence of the trade contractor or supplier: (i) carrying out or failing to "
     "carry out the trade works."),
    ("4.", "PAYMENT",
     "The principal contractor must pay the trade contractor or supplier the trade contract "
     "sum within 30 days of the later of: (a) receipt of a valid tax invoice; and (b) "
     "completion of the trade works to the reasonable satisfaction of the principal "
     "contractor."),
    ("5.", "VARIATIONS",
     "The principal contractor may direct the trade contractor or supplier to vary the trade "
     "works. The trade contractor or supplier must not vary the trade works without the "
     "principal contractor's written direction. The trade contractor or supplier must, "
     "within 5 business days of receiving a direction, provide a price for the variation."),
    ("6.", "SAFETY",
     "The trade contractor or supplier must comply with all applicable work health and safety "
     "laws and any reasonable direction of the principal contractor relating to safety. "
     "The trade contractor or supplier must provide a Safe Work Method Statement (SWMS) "
     "before commencing any high-risk construction work."),
    ("7.", "DEFECTS",
     "The trade contractor or supplier must rectify any defects in the trade works notified "
     "by the principal contractor within 7 days of notification. If the trade contractor or "
     "supplier fails to rectify defects within the specified time, the principal contractor "
     "may rectify the defects and deduct the cost from amounts owing."),
    ("8.", "TIME",
     "The trade contractor or supplier must commence the trade works on the date directed by "
     "the principal contractor and complete the trade works by the date specified in this "
     "Purchase Order. Time is of the essence."),
    ("9.", "WARRANTIES",
     "The trade contractor or supplier warrants that: (a) The trade works will be carried "
     "out in a proper and skillful manner; (b) Materials supplied by it will be suitable, "
     "new and free of defects; and (c) It holds all licenses required to carry out the "
     "trade works."),
    ("10.", "DEFAULT BY SUB-CONTRACTOR",
     "If the Subcontractor: (a) fails to proceed with the Works with reasonable diligence "
     "or in a competent manner; (b) Refuses to comply with a written or verbal direction "
     "from the Builder to remedy any defective or improper work; (c) otherwise, be guilty "
     "of substantial breach of the provisions of the Subcontract. "
     "The builder may give written or oral notice to the Subcontractor terminating the "
     "Work Order immediately. If the Builder elects to terminate the Work Order pursuant "
     "to this clause the Builder may: (e) complete the Works; and (f) take possession of "
     "and use and permit other persons to use any materials, construction plant and other "
     "things which are owned by the Subcontractor as may be necessary for the purpose of "
     "completing the Works; And in any event shall be entitled to damages from the "
     "Subcontractor for any loss suffered."),
]


# =============================================================================
# Styles
# =============================================================================

FONT_TITLE = Font(name="Arial", size=14, bold=True)
FONT_HEADING = Font(name="Arial", size=12, bold=True)
FONT_NORMAL = Font(name="Arial", size=10)
FONT_BOLD = Font(name="Arial", size=10, bold=True)
FONT_SMALL = Font(name="Arial", size=8)
FONT_SMALL_BOLD = Font(name="Arial", size=8, bold=True)
FONT_ITEM = Font(name="Arial", size=9)
FONT_ITEM_BOLD = Font(name="Arial", size=9, bold=True)

THIN = Side(style="thin")
MEDIUM = Side(style="medium")
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_BOTTOM_MEDIUM = Border(bottom=MEDIUM)
BORDER_BOTTOM_THIN = Border(bottom=THIN)

FILL_LIGHT_GRAY = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

FMT_CURRENCY = '$#,##0.00'
FMT_QTY = '#,##0.00'
FMT_DATE = 'DD MMM YY'


# =============================================================================
# Helpers
# =============================================================================

def _set_col_widths(ws, widths: dict[int, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def _generate_creditor_code(vendor_name: str) -> str:
    """Generate a 6-char creditor code from vendor name.
    Pattern: first 3 letters (uppercase) + 3-digit sequential number.
    """
    alpha = ''.join(c for c in vendor_name.upper() if c.isalpha())[:3]
    # Use a simple hash of the full name for determinism
    seq = sum(ord(c) for c in vendor_name) % 1000
    return f"{alpha}{seq:03d}"


def _format_date(d: str | None) -> str:
    """Format a date string as 'DD MMM YY' uppercase."""
    if not d:
        return datetime.now().strftime("%d %b %y").upper()
    try:
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%dT%H:%M:%S"):
            try:
                return datetime.strptime(d[:10], fmt[:10] if len(fmt) > 10 else fmt).strftime("%d %b %y").upper()
            except ValueError:
                continue
    except Exception:
        pass
    return d


# =============================================================================
# Border application — matches example xlsx cell-by-cell
# =============================================================================

def _apply_po_borders(ws, item_count: int) -> None:
    """Apply borders to the PO sheet, matching the example xlsx exactly.

    Rules extracted from the reference xlsx at data/PO_examples/example_PO_xlsx.xlsx:
      - Row 1 (A-H): bottom thin (separator under PURCHASING OFFICER)
      - Row 8 (A-H): top thin + bottom thin (title box)
      - Rows 9-15 col C: left+right thin (order-info text box); row 11 left only (C11:D11 merged)
      - Rows 9-15 col F: left thin (vertical divider for order-info values)
      - Row 9 cols G-H: top thin; Row 15 (A-H): bottom thin (separator above scope)
      - Row 25 (A-H): top medium + bottom medium (table header)
      - First data row 26: medium top on description cols (B-C)
      - Item rows (A-H): all thin borders
      - Totals 29-31 (G-H): top medium on first, bottom medium on last, thin sides
      - Footer rows 37-45: outer medium box (A-left, H-right, 37-top, 45-bottom)
        with internal divider at column D (right thin)
    """
    # ── Row 1: bottom thin separator ────────────────────────────────
    for col in range(1, 9):
        _set_border(ws, 1, col, bottom="thin")

    # ── Row 8: title box (A-H, top thin + bottom thin) ──────────────
    for col in range(1, 9):
        _set_border(ws, 8, col, top="thin", bottom="thin")

    # ── Rows 9-15: order-info text box (col C) + value divider (col F) ─
    for row in range(9, 16):
        if row == 11:
            # C11:D11 merged — right border absorbed; only left on C
            _set_border(ws, row, 3, left="thin")
        else:
            _set_border(ws, row, 3, left="thin", right="thin")
        # Col F vertical divider (order-info values on its left)
        _set_border(ws, row, 6, left="thin")
    # Top of the box (row 9): C, F, G, H
    _set_border(ws, 9, 3, top="thin")
    _set_border(ws, 9, 6, top="thin")
    _set_border(ws, 9, 7, top="thin")
    _set_border(ws, 9, 8, top="thin")
    # Bottom of the box (row 15): all cols
    for col in range(1, 9):
        _set_border(ws, 15, col, bottom="thin")

    # ── Row 25 (A-H): table header ──────────────────────────────────
    for col in range(1, 9):
        _set_border(ws, 25, col, top="medium", bottom="medium")
    for col in [1, 2, 4, 5, 6, 7, 8]:
        _set_border(ws, 25, col, left="thin", right="thin")
    _set_border(ws, 25, 3, right="thin")  # C: right thin (part of B-C merge)

    # ── Data rows: thin all around ──────────────────────────────────
    last_item_row = 25 + item_count  # row 27 for 2 items
    for row in range(26, last_item_row + 1):
        for col in [1, 2, 4, 5, 6, 7, 8]:
            _set_border(ws, row, col, left="thin", right="thin", top="thin", bottom="thin")
        # Col 3 (part of B-C merge)
        _set_border(ws, row, 3, right="thin", top="thin", bottom="thin")

    # First item row (26) gets medium top on description columns
    _set_border(ws, 26, 2, top="medium")
    _set_border(ws, 26, 3, top="medium")

    # ── Totals (rows 29-31, cols G-H) ───────────────────────────────
    _set_border(ws, 29, 7, left="thin", right="thin", top="medium", bottom="thin")
    _set_border(ws, 29, 8, left="thin", right="thin", top="medium", bottom="thin")
    _set_border(ws, 30, 7, left="thin", right="thin", top="thin", bottom="thin")
    _set_border(ws, 30, 8, left="thin", right="thin", top="thin", bottom="thin")
    _set_border(ws, 31, 7, left="thin", right="thin", top="thin", bottom="medium")
    _set_border(ws, 31, 8, left="thin", right="thin", top="thin", bottom="medium")

    # ── Footer box (rows 37-45) ─────────────────────────────────────
    # Outer frame: left medium on A, right medium on H, divider right thin at col D
    for row in range(37, 46):
        _set_border(ws, row, 1, left="medium")
        _set_border(ws, row, 8, right="medium")
        _set_border(ws, row, 4, right="thin")  # internal divider at col D
    # Top & bottom
    for col in range(1, 9):
        _set_border(ws, 37, col, top="medium")
        _set_border(ws, 45, col, bottom="medium")


def _set_border(ws, row: int, col: int, left=None, right=None, top=None, bottom=None):
    """Merge new border sides onto an existing cell border."""
    cell = ws.cell(row=row, column=col)
    existing = cell.border
    new_left = Side(style=left) if left else existing.left
    new_right = Side(style=right) if right else existing.right
    new_top = Side(style=top) if top else existing.top
    new_bottom = Side(style=bottom) if bottom else existing.bottom
    cell.border = Border(left=new_left, right=new_right, top=new_top, bottom=new_bottom)


# =============================================================================
# PO Document Generator
# =============================================================================


def generate_po_document(commitment: dict, output_dir: Path) -> Path:
    """Generate a Purchase Order matching the exact Welink xlsx template.

    Layout verified against the user's example PO xlsx at
    data/PO_examples/example_PO_xlsx.xlsx — every cell position,
    merge, font size, and alignment is matched to that reference.

    10-column portrait layout (A-J):
      - Rows 1-2:     empty (spacers)
      - Row 3:        logo area (height 18)
      - Rows 4-6:     company name + address + ABN (start at col D)
      - Row 7:        spacer (height 11.5)
      - Row 8:        "Purchase Order" title, merged A-H, 16pt bold
      - Rows 9-15:    Supplier (left A-B) + Order info (right C-D)
      - Row 16:       spacer
      - Rows 17-18:   Scope statement
      - Row 19:       Project row (labels A + F, values C-E + G)
      - Row 20:       Delivery row
      - Row 23:       Attention row
      - Row 25:       Items table header
      - Rows 26-27:   Item data rows
      - Rows 29-31:   Totals (G-H, right-aligned)
      - Rows 33-34:   Special Instructions
      - Rows 37-43:   Footer (PURCHASING OFFICER, Requested by, Approved by)
    """
    wb = openpyxl.Workbook()

    # ── Sheet 1: Purchase Order ─────────────────────────────────────
    ws = wb.active
    ws.title = "Purchase Order"

    # Page setup — A4 portrait
    ws.page_setup.paperSize = 9        # A4
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.6
    ws.page_margins.bottom = 0.6

    # Column widths — exact match to example xlsx
    _set_col_widths(ws, {
        1: 11.36,   # A: Supplier/Project labels, Item #
        2: 17.18,   # B: vendor name, description
        3: 11.36,   # C: order-info text, project value (merged B-C for desc)
        4: 9.63,    # D: Qty
        5: 8.54,    # E: Unit
        6: 10.36,   # F: Rate, order-info values (Order No, Date, Creditor)
        7: 11.36,   # G: Discount, Project Code value, totals label, footer right
        8: 13.0,    # H: Amount, totals value
    })

    # ── Extract data ────────────────────────────────────────────────
    ref = commitment.get("reference_number", "PO-XXXX")
    vendor = commitment.get("vendor_name", "Vendor")
    vendor_address = commitment.get("vendor_address", "")
    vendor_abn = commitment.get("vendor_abn", "")
    project_name = commitment.get("project_name", "Project")
    project_code = commitment.get("project_code", "")
    project_location = commitment.get("project_location", "")
    delivery_date = commitment.get("delivery_date", "")
    delivery_instructions = commitment.get("delivery_instructions", "")
    attention = commitment.get("attention", "")
    special_instructions = commitment.get("special_instructions", "")
    approved_by = commitment.get("approved_by", "ACHEN")

    # Parse vendor address into max 2 lines
    addr_lines = [l.strip() for l in vendor_address.split(",")] if vendor_address else []
    if len(addr_lines) >= 2:
        addr1 = addr_lines[0]
        addr2 = ", ".join(addr_lines[1:]).strip()
    elif len(addr_lines) == 1:
        addr1 = addr_lines[0]
        addr2 = ""
    else:
        addr1 = ""
        addr2 = ""

    po_date = _format_date(None)
    creditor_code = _generate_creditor_code(vendor)
    po_num = ref.replace("PO", "") if ref.startswith("PO") else ref
    order_number = f"{project_code} - {po_num}" if project_code else ref

    # ── Row heights (match example) ─────────────────────────────────
    for r, h in {3: 18, 7: 11.5, 8: 33, 9: 17.5, 10: 14.5, 11: 13.5,
                 12: 14, 13: 16, 14: 17, 15: 15, 16: 19.5, 19: 21.5,
                 20: 14.5, 23: 14, 25: 14.5, 28: 14.5, 33: 14.5, 36: 14.5,
                 41: 12.5, 42: 12.5, 43: 15, 44: 15, 45: 14.5}.items():
        ws.row_dimensions[r].height = h

    # ── Row 3: Logo ────────────────────────────────────────────────
    if LOGO_PATH.exists():
        try:
            logo = openpyxl.drawing.image.Image(str(LOGO_PATH))
            logo.width = 180
            logo.height = 52
            logo.anchor = "A3"
            ws.add_image(logo)
        except Exception:
            pass

    # ── Rows 3-6: Company details (column D) ───────────────────────
    ws.cell(row=3, column=4, value=WELINK_NAME).font = Font(name="Arial", size=14, bold=True)
    ws.cell(row=4, column=4, value=WELINK_ADDRESS).font = FONT_NORMAL
    ws.cell(row=5, column=4, value=WELINK_SUBURB).font = FONT_NORMAL
    ws.cell(row=6, column=4, value=f"ABN: {WELINK_ABN}").font = FONT_NORMAL

    # ── Row 8: "Purchase Order" title ──────────────────────────────
    ws.merge_cells("A8:H8")
    c = ws.cell(row=8, column=1, value="Purchase Order")
    c.font = Font(name="Arial", size=16, bold=True)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── Rows 9-15: Supplier (left A-B) + Order info (C text, F values) ─
    # Row 9
    ws.cell(row=9, column=1, value="Supplier").font = FONT_BOLD
    ws.cell(row=9, column=3, value="Order Number must be quoted on").font = FONT_NORMAL
    ws.cell(row=9, column=6, value=f"Order No: {order_number}").font = FONT_SMALL

    # Row 10
    ws.merge_cells("A10:B10")
    ws.cell(row=10, column=1, value=vendor).font = FONT_NORMAL
    ws.cell(row=10, column=1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.cell(row=10, column=3, value="Delivery Dockets and Invoices.").font = FONT_NORMAL
    ws.cell(row=10, column=6, value=f"Date: {po_date}").font = FONT_SMALL

    # Row 11: vendor address line 1
    ws.merge_cells("A11:B11")
    ws.cell(row=11, column=1, value=addr1).font = FONT_NORMAL
    ws.cell(row=11, column=1).alignment = ALIGN_LEFT
    ws.merge_cells("C11:D11")  # empty merge (spacing in order-info box)
    ws.cell(row=11, column=6, value="Creditor Phone:").font = FONT_SMALL

    # Row 12: vendor address line 2 — merged down to row 13 (A12:B13) so a
    # long address wraps and displays onto the next row (A13), top-left aligned.
    ws.merge_cells("A12:B13")
    ws.cell(row=12, column=1, value=addr2).font = FONT_NORMAL
    ws.cell(row=12, column=1).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.cell(row=12, column=3, value="All Enquiries and Correspondence").font = FONT_NORMAL
    ws.cell(row=12, column=6, value=f"Creditor Code: {creditor_code}").font = FONT_SMALL

    # Row 13
    ws.cell(row=13, column=3, value="must be addressed to:").font = FONT_NORMAL

    # Row 14: vendor ABN + email
    ws.merge_cells("A14:B14")
    ws.cell(row=14, column=1, value=f"ABN: {vendor_abn}" if vendor_abn else "").font = FONT_NORMAL
    ws.cell(row=14, column=1).alignment = ALIGN_LEFT
    ws.cell(row=14, column=3, value=f"✉ {WELINK_EMAIL}").font = FONT_NORMAL

    # Row 15: contact phone
    ws.cell(row=15, column=3, value=f"☎ {WELINK_PHONE}").font = FONT_NORMAL

    # Row 16: spacer
    ws.row_dimensions[16].height = 19.5

    # ── Rows 17-18: Scope statement (split across two rows) ─────────
    ws.merge_cells("A17:H17")
    ws.cell(row=17, column=1,
            value="Please supply the following goods and/or services in accordance "
                  "with the Terms and Conditions of ").font = FONT_NORMAL
    ws.cell(row=18, column=1,
            value="this Purchase Order and annexed documentation. Prices are exclusive of GST").font = FONT_NORMAL

    # ── Row 19: Project ────────────────────────────────────────────
    ws.cell(row=19, column=1, value="Project:").font = FONT_BOLD
    ws.merge_cells("C19:E19")
    project_full = f"{project_name} — {project_location}" if project_location else project_name
    ws.cell(row=19, column=3, value=project_full).font = FONT_NORMAL
    ws.cell(row=19, column=6, value="Project Code:").font = FONT_BOLD
    ws.cell(row=19, column=7, value=project_code).font = FONT_BOLD

    # ── Row 20: Delivery ───────────────────────────────────────────
    ws.cell(row=20, column=1, value="Delivery Instructions:").font = FONT_BOLD
    ws.merge_cells("C20:E20")
    ws.cell(row=20, column=3, value=delivery_instructions).font = FONT_NORMAL
    ws.cell(row=20, column=6, value="Delivery date required:").font = FONT_BOLD
    ws.cell(row=20, column=7, value=_format_date(delivery_date) if delivery_date else "").font = FONT_NORMAL

    # ── Row 23: Attention ──────────────────────────────────────────
    ws.cell(row=23, column=1, value="Attention:").font = FONT_BOLD
    ws.merge_cells("C23:E23")
    ws.cell(row=23, column=3, value=attention).font = FONT_NORMAL

    # ── Row 25: Items table header ─────────────────────────────────
    ws.merge_cells("B25:C25")
    headers = [
        (1, "Item"), (2, "Description"), (4, "Qty"), (5, "Unit"),
        (6, "Rate"), (7, "Discount"), (8, "Amount"),
    ]
    for col, h in headers:
        c = ws.cell(row=25, column=col, value=h)
        c.font = FONT_ITEM_BOLD
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── Rows 26+: Item data ────────────────────────────────────────
    items = commitment.get("items", [])
    total_ex = 0.0
    item_row = 26

    for i, item in enumerate(items):
        unit = item.get("unit", "item")
        is_note = (unit.upper() == "NOTE")

        qty = item.get("qty", 0) if not is_note else 0
        rate = item.get("rate", 0) if not is_note else 0
        amount = item.get("amount", qty * rate) if not is_note else 0
        if not is_note:
            total_ex += amount

        # Col A: item number
        c = ws.cell(row=item_row, column=1, value=i + 1)
        c.font = FONT_ITEM
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Col B-C: description (merged)
        ws.merge_cells(start_row=item_row, start_column=2, end_row=item_row, end_column=3)
        c = ws.cell(row=item_row, column=2, value=item.get("description", ""))
        c.font = FONT_ITEM
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Col D: qty
        c = ws.cell(row=item_row, column=4, value="N/A" if is_note else qty)
        c.font = FONT_ITEM
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if not is_note:
            c.number_format = FMT_QTY

        # Col E: unit
        c = ws.cell(row=item_row, column=5, value=unit)
        c.font = FONT_ITEM
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Col F: rate
        c = ws.cell(row=item_row, column=6, value=rate)
        c.font = FONT_ITEM
        c.alignment = Alignment(horizontal="right", vertical="center")
        if not is_note:
            c.number_format = FMT_CURRENCY

        # Col G: discount
        c = ws.cell(row=item_row, column=7, value=0)
        c.font = FONT_ITEM
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.number_format = FMT_CURRENCY

        # Col H: amount
        c = ws.cell(row=item_row, column=8, value=amount)
        c.font = FONT_ITEM
        c.alignment = Alignment(horizontal="right", vertical="center")
        if not is_note:
            c.number_format = FMT_CURRENCY

        item_row += 1

    # ── Totals (rows 29-31, columns G-H) ───────────────────────────
    gst = round(total_ex * 0.10, 2)
    gross = total_ex + gst

    for i, (label, val) in enumerate([("Total $", total_ex), ("GST", gst), ("Gross", gross)]):
        r = 29 + i
        ws.cell(row=r, column=7, value=label).font = FONT_BOLD
        ws.cell(row=r, column=7).alignment = Alignment(horizontal="right", vertical="center")
        c = ws.cell(row=r, column=8, value=val)
        c.font = FONT_BOLD
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.number_format = FMT_CURRENCY

    # ── Special Instructions (rows 33-34) ──────────────────────────
    if special_instructions:
        ws.cell(row=33, column=1, value="Special Instructions:").font = FONT_BOLD
        ws.merge_cells("A34:H36")
        ws.cell(row=34, column=1, value=special_instructions).font = FONT_NORMAL
        ws.cell(row=34, column=1).alignment = ALIGN_LEFT_TOP

    # ── Footer (rows 37-43) ────────────────────────────────────────
    ws.cell(row=37, column=5, value="PURCHASING OFFICER").font = FONT_NORMAL
    ws.cell(row=39, column=6, value="." * 47).font = FONT_NORMAL
    ws.cell(row=41, column=5, value="Requested by").font = FONT_NORMAL
    ws.cell(row=41, column=7, value="<<REQ>>").font = FONT_NORMAL
    ws.cell(row=43, column=5, value="Approved by ").font = FONT_NORMAL
    ws.cell(row=43, column=7, value=approved_by).font = FONT_NORMAL

    # ── Borders ────────────────────────────────────────────────────
    _apply_po_borders(ws, len(items))

    # ── Sheet 2: Terms & Conditions (starts on a new page) ──────────
    ws2 = wb.create_sheet("Terms & Conditions")
    ws2.column_dimensions["A"].width = 6
    ws2.column_dimensions["B"].width = 85
    # Total A+B = 91 units ≈ 8.13". Zero margins → full page width.
    # Excel's printer-minimum margin (~0.2") → printable ≈ 7.87".
    # FitToWidth=1 scales ~3% to fit — mild, clean render at this ratio.
    ws2.page_setup.paperSize = 9
    ws2.page_setup.orientation = "portrait"
    ws2.page_setup.fitToWidth = 1   # column fits 1 page at 100% → no actual scaling
    ws2.page_setup.fitToHeight = 0
    ws2.page_margins.top = 0.4      # print-title row sits within top margin
    ws2.page_margins.bottom = 0.4   # footer: page number centered
    ws2.page_margins.left = 0.4
    ws2.page_margins.right = 0.4    # balanced margins

    # Page number in the footer, centered at the bottom of each page
    ws2.oddFooter.center.text = "Page &P of &N"
    ws2.oddFooter.center.size = 8
    ws2.oddFooter.center.font = "Arial"

    # ── Row 1: "Continued" print-title row (repeats at top of every page) ─
    # Built as a body row instead of a page-header so we can apply a thick
    # bottom border and control exact positioning.
    ws2.merge_cells("A1:B1")
    c = ws2.cell(row=1, column=1, value=f"Purchase Order {order_number}  Continued")
    c.font = Font(name="Arial", size=10, bold=True)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[1].height = 20
    # Thick bottom border under the "Continued" row — visual separator
    _tcell = ws2.cell(row=1, column=1)
    _tcell.border = Border(bottom=Side(style="medium"))
    _tcell = ws2.cell(row=1, column=2)
    _tcell.border = Border(bottom=Side(style="medium"))
    # Set as print title so it repeats on every page of this sheet
    ws2.print_title_rows = "1:1"

    # Spacer
    ws2.row_dimensions[2].height = 4

    # T&C title (now starts at row 3)
    ws2.merge_cells("A3:B3")
    ws2.cell(row=3, column=1, value="PURCHASE ORDER TERMS & CONDITIONS").font = Font(name="Arial", size=14, bold=True)
    ws2.cell(row=3, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[3].height = 22
    ws2.row_dimensions[4].height = 4

    import math
    # Col B (85) at Arial 10 ≈ 115 chars/line
    CHARS_PER_LINE2 = 115
    PTS_PER_LINE2 = 14  # extra safety margin against text overlap between clauses

    r = 5
    for clause_num, clause_title, clause_body in PO_TERMS:
        # Clause title: "1.  TRADE WORKS" — Arial 11 bold
        ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        ws2.cell(row=r, column=1, value=f"{clause_num}  {clause_title}").font = Font(name="Arial", size=11, bold=True)
        ws2.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="top")
        ws2.row_dimensions[r].height = 15
        r += 1

        # Body — Arial 10, wrapped, left, top
        ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        ws2.cell(row=r, column=1, value=clause_body).font = Font(name="Arial", size=10)
        ws2.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
        lines = 0
        for seg in str(clause_body).split("\n"):
            lines += max(1, math.ceil(len(seg) / CHARS_PER_LINE2))
        ws2.row_dimensions[r].height = max(14, lines * PTS_PER_LINE2 + 4)
        r += 1

        # Minimal spacer between clauses
        ws2.row_dimensions[r].height = 3
        r += 1

    # Bottom marker
    ws2.cell(row=r, column=1, value=f"Purchase Order {order_number}  Continued").font = FONT_SMALL
    ws2.cell(row=r, column=1).alignment = Alignment(horizontal="center")
    ws2.row_dimensions[r].height = 14

    # ── Save ────────────────────────────────────────────────────────
    safe_project = (project_name or "Project").replace(" ", "_")
    safe_ref = ref.replace("/", "-").replace(" ", "_")
    safe_vendor = vendor.replace("/", "-").replace(" ", "_").replace(",", "")
    filename = f"{safe_ref} - {safe_vendor} - {safe_project}.xlsx"
    output_path = output_dir / filename
    wb.save(output_path)
    wb.close()

    return output_path


# =============================================================================
# Subcontract Document Generator
# =============================================================================


def generate_subcontract_document(commitment: dict, output_dir: Path) -> Path:
    """Generate a Subcontract agreement Excel document from commitment data.

    This is a comprehensive subcontract document with scope, schedule of rates,
    retention, insurance, and contract terms. For a formal subcontract, a Word
    (.docx) document is typically the legal instrument — the Excel serves as
    the commercial schedule.

    Args:
        commitment: Full commitment dict with items and vendor info
        output_dir: Directory to save the output file

    Returns:
        Path to the generated xlsx file
    """
    wb = openpyxl.Workbook()

    # ── Sheet 1: Subcontract Summary ─────────────────────────────────
    ws = wb.active
    ws.title = "Subcontract"

    ref = commitment.get("reference_number", "SXX")
    vendor_name = commitment.get("vendor_name", "Subcontractor")
    project_name = commitment.get("project_name", "Project")
    date_str = datetime.now().strftime("%d/%m/%Y")
    total = commitment.get("commitment_value", 0)

    # Header
    ws.merge_cells("A1:G1")
    ws.cell(row=1, column=1, value="Welink Construction Pty Ltd").font = TITLE_FONT

    ws.merge_cells("A2:G2")
    ws.cell(row=2, column=1, value="SUBCONTRACT AGREEMENT").font = HEADER_FONT
    ws.cell(row=2, column=1).alignment = CENTER_ALIGN

    # Subcontract Details
    row = 4
    details = [
        ("Subcontract No:", ref),
        ("Date:", date_str),
        ("Project:", project_name),
        ("Subcontractor:", vendor_name),
        ("Subcontractor Type:", commitment.get("vendor_type", "").title()),
        ("Head Contract:", f"AS 4000-1997 (Lump Sum)"),
        ("Subcontract Sum (ex GST):", total),
        ("Retention:", f"{commitment.get('retention_pct', 5)}%"),
    ]
    if commitment.get("start_date"):
        details.append(("Commencement:", commitment["start_date"]))
    if commitment.get("end_date"):
        details.append(("Completion:", commitment["end_date"]))

    for label, value in details:
        ws.cell(row=row, column=1, value=label).font = BOLD_FONT
        ws.cell(row=row, column=1).alignment = RIGHT_ALIGN
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = NORMAL_FONT
        if isinstance(value, (int, float)) and value > 0 and label.startswith("Subcontract Sum"):
            cell.number_format = CURRENCY_FORMAT
        row += 1

    # Scope
    row += 1
    ws.cell(row=row, column=1, value="Scope of Works:").font = BOLD_FONT
    ws.merge_cells(start_row=row, start_column=2, end_row=row + 2, end_column=7)
    ws.cell(row=row, column=2, value=commitment.get("description", "")).font = NORMAL_FONT
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    row += 3

    # Schedule of Rates Table
    row += 1
    ws.cell(row=row, column=1, value="SCHEDULE OF RATES").font = HEADER_FONT
    row += 1

    headers = ["#", "Description", "Qty", "Unit", "Rate ($)", "Amount ($)"]
    col_widths = {1: 6, 2: 45, 3: 10, 4: 8, 5: 14, 6: 16}
    _set_col_widths(ws, col_widths)

    for col, header in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=header)
    _style_header_row(ws, row, len(headers))

    row += 1
    items = commitment.get("items", [])
    for item in items:
        qty = item.get("qty", 0)
        rate = item.get("rate", 0)
        amount = item.get("amount", qty * rate)

        ws.cell(row=row, column=1, value=item.get("item_number", ""))
        ws.cell(row=row, column=2, value=item.get("description", ""))
        ws.cell(row=row, column=3, value=qty)
        ws.cell(row=row, column=4, value=item.get("unit", "item"))
        ws.cell(row=row, column=5, value=rate)
        ws.cell(row=row, column=6, value=amount)

        for col in range(1, len(headers) + 1):
            _style_data_cell(ws, row, col, is_currency=col in (5, 6))
        ws.cell(row=row, column=1).alignment = CENTER_ALIGN
        row += 1

    # Totals
    for label, fill_color in [("SUBTOTAL (ex GST)", TOTAL_FILL),
                               ("GST (10%)", TOTAL_FILL),
                               ("SUBCONTRACT SUM (incl GST)", TOTAL_FILL)]:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.cell(row=row, column=1, value=label).font = BOLD_FONT
        ws.cell(row=row, column=1).alignment = RIGHT_ALIGN
        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = fill_color
            ws.cell(row=row, column=col).border = THIN_BORDER

        if "GST" in label and "incl" not in label:
            val = round(total * 0.10, 2)
        elif "incl" in label:
            val = round(total * 1.10, 2)
        else:
            val = total

        ws.cell(row=row, column=6, value=val).font = BOLD_FONT
        ws.cell(row=row, column=6).number_format = CURRENCY_FORMAT
        ws.cell(row=row, column=6).alignment = RIGHT_ALIGN
        row += 1

    # ── Sheet 2: Contract Terms ──────────────────────────────────────
    ws2 = wb.create_sheet("Contract Terms")

    row = 1
    ws2.merge_cells("A1:D1")
    ws2.cell(row=1, column=1, value="SUBCONTRACT TERMS & CONDITIONS").font = HEADER_FONT
    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["B"].width = 60
    row = 3

    terms = [
        ("1.", "GENERAL", "This Subcontract is made under and shall be read with the head contract AS 4000-1997 between Welink Construction Pty Ltd and the Client. Where any inconsistency exists, the head contract shall prevail."),
        ("2.", "SCOPE", "The Subcontractor shall execute the works described in the Schedule of Rates in accordance with the drawings, specifications, and all relevant Australian Standards."),
        ("3.", "SUBCONTRACT SUM", f"The Subcontract Sum is ${total:,.2f} (ex GST). The Subcontract Sum may only be varied by a formal variation order signed by both parties."),
        ("4.", "PAYMENT", "The Subcontractor shall submit progress claims monthly. The Contractor shall assess and certify payment within 14 days. Payment shall be made within 30 days of certification."),
        ("5.", "RETENTION", f"Retention of {commitment.get('retention_pct', 5)}% shall apply to all progress payments. 50% of retention shall be released at Practical Completion. The balance shall be released at expiry of the Defects Liability Period."),
        ("6.", "INSURANCE", "The Subcontractor shall maintain Public Liability Insurance ($20M), Workers Compensation, and Professional Indemnity Insurance as applicable. Certificates of currency must be provided before works commence."),
        ("7.", "SECURITY OF PAYMENT", "This Subcontract is subject to the Building and Construction Industry Security of Payment Act. Payment claims must be endorsed accordingly."),
        ("8.", "DEFECTS LIABILITY", f"The Defects Liability Period is 12 months from Practical Completion, expiring {commitment.get('defects_liability_end', 'TBD')}."),
        ("9.", "VARIATIONS", "No variation work shall proceed without a signed variation order. The Subcontractor shall provide pricing within 5 business days of request."),
        ("10.", "TERMINATION", "The Contractor may terminate this Subcontract for default, insolvency, or failure to proceed with due diligence. 14 days' written notice is required."),
    ]

    for clause, title, body in terms:
        ws2.cell(row=row, column=1, value=clause).font = BOLD_FONT
        ws2.cell(row=row, column=2, value=title).font = BOLD_FONT
        row += 1
        ws2.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws2.cell(row=row, column=2, value=body).font = NORMAL_FONT
        ws2.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
        row += 2

    # ── Signatures ────────────────────────────────────────────────────
    row += 1
    ws2.cell(row=row, column=1, value="EXECUTED as an agreement:").font = BOLD_FONT
    row += 2
    ws2.cell(row=row, column=1, value="For and on behalf of").font = NORMAL_FONT
    ws2.cell(row=row, column=3, value="For and on behalf of").font = NORMAL_FONT
    row += 1
    ws2.cell(row=row, column=1, value="Welink Construction Pty Ltd").font = BOLD_FONT
    ws2.cell(row=row, column=3, value=vendor_name).font = BOLD_FONT
    row += 3
    ws2.cell(row=row, column=1, value="Signature: ____________________").font = NORMAL_FONT
    ws2.cell(row=row, column=3, value="Signature: ____________________").font = NORMAL_FONT
    row += 1
    ws2.cell(row=row, column=1, value="Name: ________________________").font = NORMAL_FONT
    ws2.cell(row=row, column=3, value="Name: ________________________").font = NORMAL_FONT
    row += 1
    ws2.cell(row=row, column=1, value=f"Date: {date_str}").font = NORMAL_FONT
    ws2.cell(row=row, column=3, value="Date: _______________").font = NORMAL_FONT

    # ── Save ──────────────────────────────────────────────────────────
    project = (project_name or "Project").replace(" ", "_")
    safe_ref = ref.replace("/", "-").replace(" ", "_")
    filename = f"{date_str.replace('/', '')} {safe_ref} - {vendor_name} - {project}.xlsx"
    output_path = output_dir / filename
    wb.save(output_path)
    wb.close()

    return output_path


# =============================================================================
# Cost Calculation (deterministic — NEVER LLM)
# =============================================================================


def calculate_commitment_costs(items: list[dict]) -> dict:
    """Run deterministic cost math on commitment items.

    Never uses LLM. Pure Python arithmetic.
    """
    total_ex_gst = 0.0
    for item in items:
        qty = item.get("qty", 0) or 0
        rate = item.get("rate", 0) or 0
        amount = item.get("amount") or (qty * rate)
        total_ex_gst += amount

    gst = round(total_ex_gst * 0.10, 2)
    total_incl_gst = round(total_ex_gst + gst, 2)

    return {
        "total_ex_gst": round(total_ex_gst, 2),
        "gst": gst,
        "total_incl_gst": total_incl_gst,
        "item_count": len(items),
    }
