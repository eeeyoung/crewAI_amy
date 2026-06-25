"""ProgressClaimService — Client Progress Claim workflow engine.

Owns all business logic for monthly progress claims:
  - Cashflow xlsx import → projects, work items, months, progress grid
  - Editable monthly % complete (the only manual input)
  - Deterministic claim generation (cumulative %, current claim, retention, GST)
  - Excel / PDF export

Follows the project service pattern:
  QObject + pyqtSignal + threading.Thread + queue.Queue

The UI layer (FastAPI routes, WebUI) is a thin consumer.

Key domain rules (from real Econolodge examples):
  - Cashflow stores per-month % complete; the Amount = COST × %.
  - Progress Claim uses CUMULATIVE % complete = sum of monthly % up to claim month.
  - TOTAL CLAIMED   = COST × cumulative %
  - PREVIOUSLY CLAIMED = prior claim's TOTAL CLAIMED for the same item (0 if first)
  - CURRENT CLAIM   = TOTAL CLAIMED − PREVIOUSLY CLAIMED
  - Future months in the cashflow are forecasts and are ignored when claiming.
"""
from __future__ import annotations

import os
import queue
import threading
import uuid
from datetime import datetime
from pathlib import Path

from PyQt6.QtCore import QObject, pyqtSignal

# Section codes used throughout the module
SECTION_CONSTRUCTION = "A"
SECTION_PROVISIONAL_SUMS = "B"
SECTION_PRELIMINARIES = "C"
SECTION_VARIATIONS = "D"
SECTION_PS_EXCLUDED = "E"

SECTION_LABELS = {
    SECTION_CONSTRUCTION: "Construction Works",
    SECTION_PROVISIONAL_SUMS: "Provisional Sums",
    SECTION_PRELIMINARIES: "Preliminaries",
    SECTION_VARIATIONS: "Variations",
    SECTION_PS_EXCLUDED: "Provisional Sums Excluded",
}

# Claimable sections (excluded PS is informational only — not part of the claim total)
CLAIMABLE_SECTIONS = (SECTION_CONSTRUCTION, SECTION_PROVISIONAL_SUMS,
                      SECTION_PRELIMINARIES, SECTION_VARIATIONS)

# ── Item and section types ────────────────────────────────────────────

ITEM_TYPE_WORK_ITEM = "work_item"
ITEM_TYPE_MARGIN = "margin"


# =============================================================================
# Helpers
# =============================================================================


def _resolve_output_dir() -> Path:
    """Progress claims output directory (Excel/PDF exports)."""
    from shared_tools.core.ipc_bridge import CREWAI_DIR
    out_dir = CREWAI_DIR / "progress_claims"
    out_dir.mkdir(parents=True, exist_ok=True)
    return out_dir


def _new_entry_id() -> str:
    return str(uuid.uuid4())


def _safe_float(value, default: float = 0.0) -> float:
    """Coerce an xlsx cell value to float; treat errors/blanks as 0."""
    if value is None:
        return default
    if isinstance(value, (int, float)):
        if isinstance(value, float) and (value != value):  # NaN
            return default
        return float(value)
    if isinstance(value, str):
        s = value.strip().replace("%", "").replace(",", "").replace("$", "")
        if not s or s.startswith("#"):  # Excel error like #REF!, #DIV/0!
            return default
        try:
            return float(s)
        except ValueError:
            return default
    return default


def _month_key_from_date(value) -> str | None:
    """Convert a header date cell (datetime) to a 'YYYY-MM' key."""
    if isinstance(value, datetime):
        return f"{value.year:04d}-{value.month:02d}"
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                d = datetime.strptime(value.strip()[:10], fmt)
                return f"{d.year:04d}-{d.month:02d}"
            except ValueError:
                continue
    return None


def _month_label_from_key(month_key: str) -> str:
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    try:
        y, m = month_key.split("-")
        return f"{months[int(m) - 1]} {y}"
    except Exception:
        return month_key


def _increment_month_key(month_key: str) -> str:
    """Return the next month key after the given 'YYYY-MM'."""
    try:
        y, m = month_key.split("-")
        y, m = int(y), int(m)
        m += 1
        if m > 12:
            m = 1
            y += 1
        return f"{y:04d}-{m:02d}"
    except Exception:
        return month_key


# =============================================================================
# ProgressClaimService
# =============================================================================


class ProgressClaimService(QObject):
    """Central service for the Client Progress Claim workflow.

    Signals (UI connects to these):
      project_created(entry_id)
      project_updated(entry_id)
      cashflow_imported(project_entry_id)
      cashflow_updated(project_entry_id)
      claim_generated(claim_entry_id)
      claim_updated(claim_entry_id)
      excel_generated(claim_entry_id, file_path)
      pdf_exported(claim_entry_id, file_path)
      error_occurred(entry_id, error_message)
      progress_update(percentage, description)
    """

    # ── Signals ───────────────────────────────────────────────────────
    project_created = pyqtSignal(str)
    project_updated = pyqtSignal(str)
    cashflow_imported = pyqtSignal(str)
    cashflow_updated = pyqtSignal(str)
    claim_generated = pyqtSignal(str)
    claim_updated = pyqtSignal(str)
    excel_generated = pyqtSignal(str, str)
    pdf_exported = pyqtSignal(str, str)
    error_occurred = pyqtSignal(str, str)
    progress_update = pyqtSignal(int, str)

    def __init__(self, parent: QObject | None = None):
        super().__init__(parent)
        self._work_queue: queue.Queue = queue.Queue()
        self._llm_semaphore = threading.Semaphore(1)
        self._running = False
        self._output_dir = _resolve_output_dir()

        from shared_tools.progress_claim.progress_claim_db import init_progress_claim_db
        init_progress_claim_db()

    # ── Lifecycle ─────────────────────────────────────────────────────

    def start(self) -> None:
        if self._running:
            return
        self._running = True
        self._worker_thread = threading.Thread(target=self._run_loop, daemon=True)
        self._worker_thread.start()

    def stop(self) -> None:
        self._running = False
        self._work_queue.put(None)

    def _run_loop(self) -> None:
        while self._running:
            try:
                task = self._work_queue.get(timeout=0.5)
                if task is None:
                    break
                action, kwargs = task
                handler = getattr(self, f"_handle_{action}", None)
                if handler:
                    try:
                        handler(**kwargs)
                    except Exception as e:
                        entry_id = kwargs.get("entry_id") or kwargs.get("project_entry_id") or ""
                        self.error_occurred.emit(str(entry_id), str(e))
            except queue.Empty:
                continue

    def _queue(self, action: str, **kwargs) -> None:
        self._work_queue.put((action, kwargs))

    # ── Projects ──────────────────────────────────────────────────────

    def create_project(self, data: dict) -> str:
        """Create a project. Returns entry_id."""
        from shared_tools.progress_claim.progress_claim_db import upsert_project
        entry_id = data.get("entry_id") or _new_entry_id()
        data["entry_id"] = entry_id
        data.setdefault("company_name", "Welink Construction")
        upsert_project(data)
        self.project_created.emit(entry_id)
        return entry_id

    def update_project(self, entry_id: str, **fields) -> bool:
        from shared_tools.progress_claim.progress_claim_db import update_project
        ok = update_project(entry_id, **fields)
        if ok:
            self.project_updated.emit(entry_id)
        return ok

    def get_project(self, entry_id: str) -> dict | None:
        from shared_tools.progress_claim.progress_claim_db import get_project
        return get_project(entry_id)

    def list_projects(self) -> list[dict]:
        from shared_tools.progress_claim.progress_claim_db import get_projects
        return get_projects()

    def delete_project(self, entry_id: str) -> bool:
        from shared_tools.progress_claim.progress_claim_db import delete_project
        return delete_project(entry_id)

    # ── Cashflow import ───────────────────────────────────────────────

    def import_cashflow(self, project_entry_id: str, xlsx_path: str,
                        auto_create_project: bool = True) -> None:
        """Queue a cashflow import. Emits cashflow_imported when done.

        If the project does not yet exist and auto_create_project is True,
        a project record is created from the xlsx header metadata.
        """
        self._queue("import_cashflow", project_entry_id=project_entry_id,
                    xlsx_path=xlsx_path, auto_create_project=auto_create_project)

    def _handle_import_cashflow(self, project_entry_id: str, xlsx_path: str,
                                auto_create_project: bool = True) -> None:
        import openpyxl
        from shared_tools.progress_claim import progress_claim_db as db

        if not Path(xlsx_path).exists():
            raise FileNotFoundError(f"Cashflow file not found: {xlsx_path}")

        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb.active or wb[wb.sheetnames[0]]

        # Resolve / create project
        project = db.get_project(project_entry_id) if project_entry_id else None
        header = _parse_cashflow_header(ws)
        if project is None:
            if not auto_create_project:
                raise ValueError(f"Project not found: {project_entry_id}")
            project_entry_id = self.create_project({
                "name": header.get("project_name", Path(xlsx_path).stem),
                "job_number": header.get("job_number", ""),
                "client": header.get("client", ""),
                "client_contact": header.get("client_contact", ""),
                "superintendent": header.get("superintendent", ""),
                "company_name": header.get("company_name", "Welink Construction"),
                "source_type": "cashflow_import",
            })
        else:
            # Update project metadata from the header
            self.update_project(project_entry_id,
                                client=header.get("client", project.get("client", "")),
                                client_contact=header.get("client_contact", project.get("client_contact", "")),
                                superintendent=header.get("superintendent", project.get("superintendent", "")))
        self.update_project(project_entry_id, cashflow_path=str(Path(xlsx_path).resolve()))

        # Clear any prior cashflow data for a clean re-import
        db.clear_cashflow_for_project(project_entry_id)

        # Detect month columns from the header rows
        month_cols = _detect_month_columns(ws)  # list of (month_key, pct_col_letter, amt_col_letter)
        month_ids: list[int] = []
        for idx, (mkey, _pct, _amt) in enumerate(month_cols):
            mid = db.upsert_month({
                "project_entry_id": project_entry_id,
                "month_key": mkey,
                "month_label": _month_label_from_key(mkey),
                "month_index": idx,
            })
            month_ids.append(mid)
        # Map month_key -> month_id for progress lookup
        month_id_by_key = {mkey: mid for (mkey, _p, _a), mid in zip(month_cols, month_ids)}

        # Walk rows, classifying into sections
        section = SECTION_CONSTRUCTION
        section_label = SECTION_LABELS[SECTION_CONSTRUCTION]
        sort_counter = 0
        item_number_in_section = 0
        progress_records: list[dict] = []
        cost_by_section: dict[str, float] = {}
        # Track sections seen (code -> (label, claimable, order, section_type))
        # to populate the cashflow_sections table.  Register the implicit
        # Construction section first (it has no header row in the file).
        sections_seen: dict[str, list] = {
            SECTION_CONSTRUCTION: [SECTION_LABELS[SECTION_CONSTRUCTION], True, 0, "normal"]
        }
        # Track which sections have a margin row (to avoid duplicating in
        # the post-import guard).
        sections_with_margin: set = set()
        section_order = 1

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            desc_cell = row[0].value if len(row) > 0 else None
            if desc_cell is None:
                continue
            desc = str(desc_cell).strip()
            if not desc:
                continue

            new_section = _classify_section_header(desc)
            if new_section is not None:
                section = new_section
                section_label = desc  # use the actual Excel cell text, not the hardcoded default
                item_number_in_section = 0
                if section not in sections_seen:
                    sec_type = "preliminary" if section == SECTION_PRELIMINARIES else "normal"
                    sections_seen[section] = [section_label, section != SECTION_PS_EXCLUDED, section_order, sec_type]
                    section_order += 1
                continue

            # Skip subtotal / total / margin / reconciliation rows
            if _is_skip_row(desc):
                continue

            # A real work-item row needs a numeric cost in column B
            cost = _safe_float(row[1].value if len(row) > 1 else None)
            if cost <= 0:
                # No cost → not a claimable line item (e.g. blank placeholder rows)
                continue

            # Detect margin rows inside Preliminary sections
            item_type = ITEM_TYPE_WORK_ITEM
            if section == SECTION_PRELIMINARIES and desc.lower().startswith("margin"):
                item_type = ITEM_TYPE_MARGIN
                sections_with_margin.add(section)

            item_number_in_section += 1
            sort_counter += 1
            work_item_id = db.upsert_work_item({
                "project_entry_id": project_entry_id,
                "section": section,
                "section_label": section_label,
                "item_number": item_number_in_section,
                "description": desc,
                "cost": cost,
                "sort_order": sort_counter,
                "item_type": item_type,
            })
            if work_item_id is None:
                continue

            cost_by_section[section] = cost_by_section.get(section, 0) + cost

            # Read each month's progress. For margin items only the amount
            # column matters; % is always empty.  For normal items prefer the
            # amount cell when present, otherwise derive from %.
            if item_type == ITEM_TYPE_MARGIN:
                for (mkey, _pct_col, amt_col) in month_cols:
                    amt = _safe_float(_cell_by_letter(row, amt_col))
                    progress_records.append({
                        "work_item_id": work_item_id,
                        "month_id": month_id_by_key[mkey],
                        "percentage": 0,
                        "amount": amt,
                    })
            else:
                for (mkey, pct_col, amt_col) in month_cols:
                    amt = _safe_float(_cell_by_letter(row, amt_col))
                    pct = _safe_float(_cell_by_letter(row, pct_col))
                    if amt != 0:
                        percentage = amt / cost if cost > 0 else 0
                        amount = amt
                    elif pct != 0:
                        percentage = pct
                        amount = cost * pct
                    else:
                        continue
                    progress_records.append({
                        "work_item_id": work_item_id,
                        "month_id": month_id_by_key[mkey],
                        "percentage": percentage,
                        "amount": amount,
                    })

        if progress_records:
            db.bulk_set_progress(progress_records)

        # Populate the cashflow_sections table (clear first for re-import).
        db.clear_sections_for_project(project_entry_id)
        # Ensure the default Construction section (no explicit header) exists.
        if SECTION_CONSTRUCTION not in sections_seen:
            sections_seen[SECTION_CONSTRUCTION] = [SECTION_LABELS[SECTION_CONSTRUCTION], True, 0, "normal"]
        for code, (label, claimable, order, sec_type) in sections_seen.items():
            db.upsert_section({
                "project_entry_id": project_entry_id,
                "section_code": code,
                "section_label": label,
                "claimable": 1 if claimable else 0,
                "sort_order": order,
                "section_type": sec_type,
            })

        # Post-import guard: every Preliminary section must have a margin
        # row at the bottom.  If the imported xlsx already had one it was
        # detected above; otherwise auto-create one so the user can start
        # entering margin amounts immediately.
        for code, (_label, _claimable, _order, sec_type) in sections_seen.items():
            if sec_type == "preliminary" and code not in sections_with_margin:
                self.add_work_item(project_entry_id, code,
                                   description="Margins (incl. Variations)",
                                   cost=0, item_type=ITEM_TYPE_MARGIN)

        # Store a contract value hint from the construction total if absent
        project = db.get_project(project_entry_id)
        if project and not project.get("base_contract_amount"):
            total_cost = sum(v for (code, (label, claimable, order, sec_type)) in sections_seen.items()
                             if claimable for v in [cost_by_section.get(code, 0)])
            db.update_project(project_entry_id, base_contract_amount=total_cost)

        wb.close()
        self.progress_update.emit(100, f"Imported cashflow: {sort_counter} items, {len(month_cols)} months")
        self.cashflow_imported.emit(project_entry_id)

    # ── Cashflow read / edit ──────────────────────────────────────────

    def get_cashflow(self, project_entry_id: str) -> dict:
        """Return the full cashflow state for the UI grid."""
        from shared_tools.progress_claim import progress_claim_db as db
        project = db.get_project(project_entry_id)
        if not project:
            return {"project": None, "sections": {}, "months": []}

        items = db.get_work_items(project_entry_id)
        months = db.get_months(project_entry_id)
        # Build progress lookup: (work_item_id, month_id) -> {percentage, amount}
        all_progress = db.get_all_progress(project_entry_id)
        prog_lookup: dict[tuple[int, int], dict] = {}
        for p in all_progress:
            prog_lookup[(p["work_item_id"], p["month_id"])] = p

        # Group items by section, attach progress grid
        sections: dict[str, list[dict]] = {}
        for it in items:
            sec = it["section"]
            row = dict(it)
            grid = []
            for m in months:
                p = prog_lookup.get((it["id"], m["id"]))
                grid.append({
                    "month_id": m["id"],
                    "month_key": m["month_key"],
                    "month_label": m["month_label"],
                    "percentage": p["percentage"] if p else 0,
                    "amount": p["amount"] if p else 0,
                })
            row["progress"] = grid
            sections.setdefault(sec, []).append(row)

        # Section totals (cost + cumulative claimed so far across all months)
        month_totals = []
        for m in months:
            mt = sum(p["amount"] for p in all_progress if p["month_id"] == m["id"])
            month_totals.append({
                "month_id": m["id"],
                "month_key": m["month_key"],
                "month_label": m["month_label"],
                "total": mt,
            })

        return {
            "project": project,
            "sections": sections,
            "section_defs": db.get_sections(project_entry_id),
            "section_labels": SECTION_LABELS,
            "months": [dict(m) for m in months],
            "month_totals": month_totals,
        }

    def update_progress(self, project_entry_id: str, work_item_id: int,
                        month_id: int, percentage: float | None = None,
                        amount: float | None = None) -> None:
        """Queue a single cell update. Either percentage or amount may be
        given; the other is derived (mutual sync). Emits cashflow_updated."""
        self._queue("update_progress", project_entry_id=project_entry_id,
                    work_item_id=work_item_id, month_id=month_id,
                    percentage=percentage, amount=amount)

    def _handle_update_progress(self, project_entry_id: str, work_item_id: int,
                                month_id: int, percentage: float | None = None,
                                amount: float | None = None) -> None:
        from shared_tools.progress_claim import progress_claim_db as db
        item = next((it for it in db.get_work_items(project_entry_id) if it["id"] == work_item_id), None)
        if not item:
            return
        cost = item.get("cost", 0)
        item_type = item.get("item_type", ITEM_TYPE_WORK_ITEM)

        if item_type == ITEM_TYPE_MARGIN:
            # Margin: store amount as-is, percentage is always 0 (no % math)
            amount = _safe_float(amount) if amount is not None else 0
            percentage = 0.0
        else:
            # Mutual sync: derive the missing value from the given one.
            if percentage is None and amount is not None:
                amount = _safe_float(amount)
                percentage = (amount / cost) if cost else 0.0
            else:
                # Clamp to a sane range
                percentage = max(min(_safe_float(percentage), 1.0), -1.0)
                amount = cost * percentage

        db.bulk_set_progress([{
            "work_item_id": work_item_id,
            "month_id": month_id,
            "percentage": percentage,
            "amount": amount,
        }])
        self.cashflow_updated.emit(project_entry_id)

    # ── Cashflow drafting: months + work items ────────────────────────

    def add_month(self, project_entry_id: str, month_key: str | None = None) -> dict | None:
        """Append a new month column. If month_key is None, use the next
        calendar month after the current last month. Returns the new month."""
        from shared_tools.progress_claim import progress_claim_db as db
        months = db.get_months(project_entry_id)
        if month_key is None:
            if months:
                month_key = _increment_month_key(months[-1]["month_key"])
            else:
                month_key = datetime.now().strftime("%Y-%m")
        # month_index = max existing + 1
        next_index = (max((m["month_index"] for m in months), default=-1) + 1)
        mid = db.upsert_month({
            "project_entry_id": project_entry_id,
            "month_key": month_key,
            "month_label": _month_label_from_key(month_key),
            "month_index": next_index,
        })
        self.cashflow_updated.emit(project_entry_id)
        return db.get_month(mid) if mid else None

    def remove_month(self, project_entry_id: str, month_id: int) -> bool:
        """Remove a month column and all progress for it."""
        from shared_tools.progress_claim import progress_claim_db as db
        ok = db.delete_month(month_id)
        if ok:
            self.cashflow_updated.emit(project_entry_id)
        return ok

    def add_work_item(self, project_entry_id: str, section: str,
                      description: str = "", cost: float = 0,
                      item_type: str = "work_item") -> dict | None:
        """Add a work item to a section.

        For Preliminary sections the margin row always stays at the bottom,
        so non-margin items are inserted *above* it.
        """
        from shared_tools.progress_claim import progress_claim_db as db
        all_items = db.get_work_items(project_entry_id)
        sec_items = [i for i in all_items if i["section"] == section]
        section_def = db.get_section(project_entry_id, section)
        sec_type = section_def.get("section_type", "normal") if section_def else "normal"

        if item_type == ITEM_TYPE_MARGIN:
            # Margin always goes at the very bottom of its section
            insert_after = max((i["sort_order"] for i in sec_items), default=0)
        elif sec_type == "preliminary":
            # Insert above the margin row (if present)
            margin_items = [i for i in sec_items if i.get("item_type") == ITEM_TYPE_MARGIN]
            if margin_items:
                margin_sort = margin_items[0]["sort_order"]
                # Insert just before the margin row, then bump the margin down
                items_before_margin = [i for i in sec_items if i["sort_order"] < margin_sort]
                insert_after = max((i["sort_order"] for i in items_before_margin), default=0)
                # Bump the margin row's sort_order so it stays at the bottom
                db.update_work_item(margin_items[0]["id"], sort_order=margin_sort + 1)
            else:
                insert_after = max((i["sort_order"] for i in sec_items), default=0)
        else:
            max_sort = max((i["sort_order"] for i in all_items), default=0)
            insert_after = max((i["sort_order"] for i in sec_items), default=max_sort)

        item_number = len(sec_items) + 1
        # Bump sort_order of all items after the insertion point (except the
        # margin row which was already bumped above).
        for i in all_items:
            if i["sort_order"] > insert_after and i.get("item_type") != ITEM_TYPE_MARGIN:
                db.update_work_item(i["id"], sort_order=i["sort_order"] + 1)

        new_id = db.upsert_work_item({
            "project_entry_id": project_entry_id,
            "section": section,
            "section_label": SECTION_LABELS.get(section, section),
            "item_number": item_number,
            "description": description,
            "cost": _safe_float(cost),
            "sort_order": insert_after + 1,
            "item_type": item_type,
        })
        self.cashflow_updated.emit(project_entry_id)
        return db.get_work_item(new_id) if new_id else None

    def update_work_item(self, item_id: int, **fields) -> bool:
        """Update a work item. When cost changes, recompute all its progress
        amounts (amount = cost × percentage) — except for margin items where
        cost and amounts are independent."""
        from shared_tools.progress_claim import progress_claim_db as db
        item = db.get_work_item(item_id)
        if not item:
            return False
        if "cost" in fields:
            fields["cost"] = _safe_float(fields["cost"])
        ok = db.update_work_item(item_id, **fields)
        if ok and "cost" in fields and item.get("item_type") != ITEM_TYPE_MARGIN:
            new_cost = fields["cost"]
            for p in db.get_progress_for_item(item_id):
                db.bulk_set_progress([{
                    "work_item_id": item_id,
                    "month_id": p["month_id"],
                    "percentage": p["percentage"],
                    "amount": new_cost * p["percentage"],
                }])
            self.cashflow_updated.emit(item["project_entry_id"])
        return ok

    def remove_work_item(self, item_id: int) -> bool:
        from shared_tools.progress_claim import progress_claim_db as db
        item = db.get_work_item(item_id)
        if not item:
            return False
        if item.get("item_type") == ITEM_TYPE_MARGIN:
            raise ValueError("Cannot remove a margin item. It is automatically managed by the section.")
        ok = db.delete_work_item(item_id)
        if ok:
            self.cashflow_updated.emit(item["project_entry_id"])
        return ok

    # ── Cashflow sections (freeform add/remove/rename) ────────────────

    def add_section(self, project_entry_id: str, label: str = "",
                    claimable: bool = True, section_type: str = "normal") -> dict | None:
        from shared_tools.progress_claim import progress_claim_db as db
        code = db.next_section_code(project_entry_id)
        order = len(db.get_sections(project_entry_id))
        sid = db.upsert_section({
            "project_entry_id": project_entry_id,
            "section_code": code,
            "section_label": label or f"Section {code}",
            "claimable": 1 if claimable else 0,
            "sort_order": order,
            "section_type": section_type,
        })
        if section_type == "preliminary":
            # Auto-create a margin row at the bottom of the new section
            self.add_work_item(project_entry_id, code,
                               description="Margins (incl. Variations)",
                               cost=0, item_type=ITEM_TYPE_MARGIN)
        self.cashflow_updated.emit(project_entry_id)
        return db.get_section(project_entry_id, code) if sid else None

    def rename_section(self, project_entry_id: str, section_code: str,
                       label: str) -> bool:
        from shared_tools.progress_claim import progress_claim_db as db
        ok = db.update_section(project_entry_id, section_code, section_label=label)
        if ok:
            # keep work_items' section_label in sync for display
            for it in db.get_work_items(project_entry_id):
                if it["section"] == section_code:
                    db.update_work_item(it["id"], section_label=label)
            self.cashflow_updated.emit(project_entry_id)
        return ok

    def set_section_claimable(self, project_entry_id: str, section_code: str,
                              claimable: bool) -> bool:
        from shared_tools.progress_claim import progress_claim_db as db
        ok = db.update_section(project_entry_id, section_code,
                               claimable=1 if claimable else 0)
        if ok:
            self.cashflow_updated.emit(project_entry_id)
        return ok

    def remove_section(self, project_entry_id: str, section_code: str) -> bool:
        from shared_tools.progress_claim import progress_claim_db as db
        ok = db.delete_section(project_entry_id, section_code)
        if ok:
            self.cashflow_updated.emit(project_entry_id)
        return ok

    # ── Claim generation ──────────────────────────────────────────────

    def generate_claim(self, project_entry_id: str, claim_month: str,
                       claim_date: str | None = None) -> str | None:
        """Generate (or regenerate) a progress claim for the given month.

        Runs synchronously and returns the claim entry_id. The cumulative %
        is the sum of all monthly % up to and including claim_month; future
        months are ignored. PREVIOUSLY CLAIMED is the cumulative progress
        through the month BEFORE claim_month, computed directly from the
        cashflow (robust to skipped months and revisions).
        """
        from shared_tools.progress_claim import progress_claim_db as db

        project = db.get_project(project_entry_id)
        if not project:
            raise ValueError(f"Project not found: {project_entry_id}")

        items = db.get_work_items(project_entry_id)
        months = db.get_months(project_entry_id)
        all_progress = db.get_all_progress(project_entry_id)

        # progress lookup: (work_item_id, month_id) -> percentage
        prog_lookup: dict[tuple[int, int], float] = {}
        amt_lookup: dict[tuple[int, int], float] = {}
        for p in all_progress:
            prog_lookup[(p["work_item_id"], p["month_id"])] = p["percentage"]
            amt_lookup[(p["work_item_id"], p["month_id"])] = p["amount"]

        # Months up to and including the claim month, and strictly before it.
        # previously_claimed is the cumulative progress through the month
        # BEFORE claim_month. It is LEARNED from stored prior claims when they
        # exist (so imported / manually-adjusted prior claims are honoured):
        # for each work item, take the most recent prior claim's total_claimed
        # for that item; fall back to the cashflow-derived cumulative when no
        # prior claim covers the item.
        claimable_months = [m for m in months if m["month_key"] <= claim_month]
        prior_months = [m for m in months if m["month_key"] < claim_month]

        # Determine claim number + previous claim
        existing = db.get_claims(project_entry_id)
        # If a claim already exists for this exact month, regenerate it (same number)
        existing_for_month = next((c for c in existing if c["claim_month"] == claim_month), None)
        if existing_for_month:
            claim_number = existing_for_month["claim_number"]
            entry_id = existing_for_month["entry_id"]
            rev_number = existing_for_month.get("rev_number", 1) + 1
        else:
            claim_number = db.next_claim_number(project_entry_id)
            entry_id = _new_entry_id()
            rev_number = 1

        # Build a lookup of the most recent prior claim's per-item
        # total_claimed (claim_month < this claim's month). Prior claims are
        # ordered by claim_month ascending so later ones overwrite earlier.
        prior_claim_total_by_item: dict[int, float] = {}
        # Most recent prior claim (for the cumulative "Less Previous" summary,
        # which is remembered project-wide so manual edits to a prior claim's
        # gross/retention flow into the next claim's Less Previous).
        most_recent_prior: dict | None = None
        for c in sorted(existing, key=lambda x: x["claim_month"]):
            if c["claim_month"] >= claim_month:
                continue
            if c["entry_id"] == entry_id:
                continue  # skip self on regenerate
            most_recent_prior = c
            for ci in db.get_claim_items(c["entry_id"]):
                wid = ci.get("work_item_id")
                if wid:
                    prior_claim_total_by_item[wid] = _safe_float(ci["total_claimed"])

        # Compute per-item claim figures
        claim_items: list[dict] = []
        # Dynamic sections from the cashflow_sections table (freeform).
        section_defs = db.get_sections(project_entry_id)
        section_codes = [s["section_code"] for s in section_defs]
        section_claimable = {s["section_code"]: bool(s["claimable"]) for s in section_defs}
        section_label = {s["section_code"]: s["section_label"] for s in section_defs}
        if not section_codes:
            # fallback to legacy fixed sections
            section_codes = list(CLAIMABLE_SECTIONS) + [SECTION_PS_EXCLUDED]
            section_claimable = {s: True for s in CLAIMABLE_SECTIONS}
            section_claimable[SECTION_PS_EXCLUDED] = False
            section_label = dict(SECTION_LABELS)
        section_cumulative = {s: 0.0 for s in section_codes}
        item_no_by_section = {s: 0 for s in section_codes}
        sort_order = 0

        for it in items:
            sec = it["section"]
            it_type = it.get("item_type", ITEM_TYPE_WORK_ITEM)

            if it_type == ITEM_TYPE_MARGIN:
                # Margin: total_claimed is the sum of monthly amounts (not
                # cost × %).  previously_claimed is the sum of amounts from
                # months strictly before the claim month.
                total_claimed = sum(amt_lookup.get((it["id"], m["id"]), 0)
                                    for m in claimable_months)
                cumulative = 0.0
                cumulative_prior_amt = sum(amt_lookup.get((it["id"], m["id"]), 0)
                                           for m in prior_months)
                if it["id"] in prior_claim_total_by_item:
                    previously_claimed = prior_claim_total_by_item[it["id"]]
                else:
                    previously_claimed = cumulative_prior_amt
            else:
                cumulative = sum(prog_lookup.get((it["id"], m["id"]), 0)
                                 for m in claimable_months)
                cumulative_prior = sum(prog_lookup.get((it["id"], m["id"]), 0)
                                       for m in prior_months)
                total_claimed = it["cost"] * cumulative
                if it["id"] in prior_claim_total_by_item:
                    # Learned from a stored prior claim (imported or generated)
                    previously_claimed = prior_claim_total_by_item[it["id"]]
                else:
                    previously_claimed = it["cost"] * cumulative_prior

            current_claim = total_claimed - previously_claimed
            balance = it["cost"] - total_claimed

            if sec in item_no_by_section:
                item_no_by_section[sec] += 1
            sort_order += 1

            claim_items.append({
                "claim_entry_id": entry_id,
                "work_item_id": it["id"],
                "section": sec,
                "item_number": item_no_by_section.get(sec, 0),
                "description": it["description"],
                "cost": it["cost"],
                "cumulative_percentage": cumulative,
                "total_claimed": total_claimed,
                "previously_claimed": previously_claimed,
                "current_claim": current_claim,
                "balance_remaining": balance,
                "sort_order": sort_order,
            })
            if sec in section_cumulative:
                section_cumulative[sec] += total_claimed

        # Summary — CUMULATIVE model.
        # Section values = cumulative claimed to date per section. Gross Claim
        # for Works Completed = sum of the claimable section cumulative values.
        # Less Previous (after retention) is remembered from the most recent
        # prior claim (prior.gross − prior.retention), so manual edits to a
        # prior claim propagate. Retention = total held to date (10% of
        # cumulative, capped at 5% of contract). Net = Gross − Less Previous −
        # Retention. section_totals_json stores per-section {label, total,
        # claimable} for the summary card + export (dynamic sections).
        import json as _json
        section_totals_json = {
            s: {"label": section_label.get(s, s),
                "total": round(section_cumulative.get(s, 0.0), 4),
                "claimable": bool(section_claimable.get(s, True))}
            for s in section_codes
        }
        gross_claim = sum(section_cumulative[s] for s in section_codes
                          if section_claimable.get(s, True))
        cumulative_claimed = gross_claim  # same thing under the cumulative model

        base_contract = _safe_float(project.get("base_contract_amount"))
        retention_pct = 10.0
        retention_max_pct = 5.0
        retention_cap = base_contract * retention_max_pct / 100.0 if base_contract > 0 else float("inf")
        retention_amount = min(cumulative_claimed * retention_pct / 100.0, retention_cap)

        # Less Previous (after retention) — learned from the most recent prior
        # claim's stored gross + retention (which may have been manually edited).
        if most_recent_prior:
            prior_gross = _safe_float(most_recent_prior.get("gross_claim"))
            prior_retention = _safe_float(most_recent_prior.get("retention_amount"))
            less_previous = max(0.0, prior_gross - prior_retention)
        else:
            less_previous = 0.0

        net_claim = gross_claim - less_previous - retention_amount
        gst_amount = net_claim * 0.10
        total_incl_gst = net_claim + gst_amount

        if claim_date is None:
            claim_date = datetime.now().strftime("%Y-%m-%d")

        claim_data = {
            "entry_id": entry_id,
            "project_entry_id": project_entry_id,
            "claim_number": claim_number,
            "claim_month": claim_month,
            "claim_date": claim_date,
            "rev_number": rev_number,
            "status": "draft",
            "retention_percentage": retention_pct,
            "retention_max_percentage": retention_max_pct,
            "gross_claim": gross_claim,
            "less_previous_claims": less_previous,
            "retention_amount": retention_amount,
            "total_retention_held": retention_amount,
            "net_claim": net_claim,
            "gst_amount": gst_amount,
            "total_including_gst": total_incl_gst,
            "section_a_total": section_cumulative.get(SECTION_CONSTRUCTION, 0.0),
            "section_b_total": section_cumulative.get(SECTION_PROVISIONAL_SUMS, 0.0),
            "section_c_total": section_cumulative.get(SECTION_PRELIMINARIES, 0.0),
            "section_d_total": section_cumulative.get(SECTION_VARIATIONS, 0.0),
            "section_e_total": section_cumulative.get(SECTION_PS_EXCLUDED, 0.0),
            "cumulative_claimed": cumulative_claimed,
        }
        db.upsert_claim(claim_data)
        db.update_claim(entry_id, section_totals_json=_json.dumps(section_totals_json))
        db.clear_claim_items(entry_id)
        db.bulk_insert_claim_items(claim_items)

        self.claim_generated.emit(entry_id)
        return entry_id

    def get_claim_summary(self, claim_entry_id: str) -> dict | None:
        from shared_tools.progress_claim import progress_claim_db as db
        claim = db.get_claim(claim_entry_id)
        if not claim:
            return None
        items = db.get_claim_items(claim_entry_id)
        project = db.get_project(claim["project_entry_id"])
        # Group items by section for the UI
        sections: dict[str, list[dict]] = {}
        for it in items:
            sections.setdefault(it["section"], []).append(it)
        return {
            "claim": claim,
            "project": project,
            "items": items,
            "sections": sections,
            "section_labels": SECTION_LABELS,
        }

    def update_claim_item(self, claim_item_id: int, **fields) -> bool:
        """Edit a claim item (Manual Mode — any field editable).

        Field-aware: editing a derived field adjusts its source so the
        mathematical relationships always hold afterwards:
            current_claim      = total_claimed - previously_claimed
            balance_remaining  = cost - total_claimed
            total_claimed      = cost × cumulative_percentage
        Editing current_claim → previously_claimed = total - current.
        Editing balance       → cost = total + balance.
        Editing cumulative_%  → total_claimed = cost × %.
        Editing total_claimed → cumulative_% = total / cost.
        Editing cost          → cumulative_% = total / cost.
        Then _recompute_claim re-derives current/balance and the summary.
        """
        from shared_tools.progress_claim import progress_claim_db as db
        item = db.get_claim_item(claim_item_id)
        if not item:
            return False

        # Coerce numeric fields
        for k in ("cost", "cumulative_percentage", "total_claimed",
                  "previously_claimed", "current_claim", "balance_remaining"):
            if k in fields:
                fields[k] = _safe_float(fields[k])

        cost = _safe_float(fields.get("cost", item["cost"]))
        total = _safe_float(fields.get("total_claimed", item["total_claimed"]))
        previously = _safe_float(fields.get("previously_claimed", item["previously_claimed"]))
        cum_pct = _safe_float(fields.get("cumulative_percentage", item["cumulative_percentage"]))

        if "current_claim" in fields:
            previously = total - _safe_float(fields["current_claim"])
            fields["previously_claimed"] = previously
        if "balance_remaining" in fields:
            cost = total + _safe_float(fields["balance_remaining"])
            fields["cost"] = cost
        if "cumulative_percentage" in fields:
            total = cost * cum_pct
            fields["total_claimed"] = total
        if "total_claimed" in fields:
            if cost > 0:
                fields["cumulative_percentage"] = total / cost
        if "cost" in fields:
            if cost > 0:
                fields["cumulative_percentage"] = total / cost

        ok = db.update_claim_item(claim_item_id, **fields)
        if ok:
            self._recompute_claim(item["claim_entry_id"])
            self.claim_updated.emit(item["claim_entry_id"])
        return ok

    def _recompute_claim(self, claim_entry_id: str) -> None:
        """Recompute derived item fields + the claim summary from stored
        claim_items. Preserves total_claimed / previously_claimed / cost as
        edited; derives current_claim and balance_remaining, then rolls up
        section totals, gross, retention, net, GST, and total."""
        from shared_tools.progress_claim import progress_claim_db as db
        claim = db.get_claim(claim_entry_id)
        if not claim:
            return
        project = db.get_project(claim["project_entry_id"]) or {}
        items = db.get_claim_items(claim_entry_id)

        section_totals = {s: 0.0 for s in CLAIMABLE_SECTIONS}
        cumulative_claimed = 0.0
        less_previous = 0.0

        for it in items:
            total_claimed = _safe_float(it["total_claimed"])
            previously = _safe_float(it["previously_claimed"])
            cost = _safe_float(it["cost"])
            current = total_claimed - previously
            balance = cost - total_claimed
            db.update_claim_item(it["id"], current_claim=current,
                                 balance_remaining=balance)
            sec = it["section"]
            if sec in section_totals:
                section_totals[sec] += current
            cumulative_claimed += total_claimed
            less_previous += previously

        gross_claim = sum(section_totals[s] for s in CLAIMABLE_SECTIONS)
        base_contract = _safe_float(project.get("base_contract_amount"))
        retention_pct = _safe_float(claim.get("retention_percentage"), 10.0) or 10.0
        retention_max_pct = _safe_float(claim.get("retention_max_percentage"), 5.0) or 5.0
        retention_cap = base_contract * retention_max_pct / 100.0 if base_contract > 0 else float("inf")
        total_retention_held = min(cumulative_claimed * retention_pct / 100.0, retention_cap)
        prior_retention = min(less_previous * retention_pct / 100.0, retention_cap)
        retention_amount = max(0.0, total_retention_held - prior_retention)
        net_claim = gross_claim - retention_amount
        gst_amount = net_claim * 0.10
        total_incl_gst = net_claim + gst_amount

        db.update_claim(claim_entry_id,
                        gross_claim=gross_claim,
                        less_previous_claims=less_previous,
                        retention_amount=retention_amount,
                        total_retention_held=total_retention_held,
                        net_claim=net_claim,
                        gst_amount=gst_amount,
                        total_including_gst=total_incl_gst,
                        section_a_total=section_totals[SECTION_CONSTRUCTION],
                        section_b_total=section_totals[SECTION_PROVISIONAL_SUMS],
                        section_c_total=section_totals[SECTION_PRELIMINARIES],
                        section_d_total=section_totals[SECTION_VARIATIONS],
                        section_e_total=section_totals.get(SECTION_PS_EXCLUDED, 0.0),
                        cumulative_claimed=cumulative_claimed)

    def update_claim_summary(self, claim_entry_id: str, **fields) -> dict:
        """Manual Mode: edit the claim summary card directly.

        Editable: claim_number (unique per project), section cumulative
        values (via section_totals dict {code: total}), less_previous_claims
        (after retention), retention_amount (total held).

        Gross = sum of claimable section cumulative values. Net = Gross −
        Less Previous − Retention. GST = Net × 10%. Total = Net + GST. These
        persist, and the next claim generated for a later month initialises
        its Less Previous from this claim's Gross − Retention, so manual edits
        propagate project-wide.
        """
        import json as _json
        from shared_tools.progress_claim import progress_claim_db as db
        claim = db.get_claim(claim_entry_id)
        if not claim:
            return {"error": "Claim not found"}

        # Claim number uniqueness
        if "claim_number" in fields and fields["claim_number"] is not None:
            try:
                new_no = int(fields["claim_number"])
            except (TypeError, ValueError):
                return {"error": "Claim number must be an integer"}
            if new_no != claim["claim_number"]:
                for c in db.get_claims(claim["project_entry_id"]):
                    if c["entry_id"] != claim_entry_id and c["claim_number"] == new_no:
                        return {"error": f"Claim No.{new_no:02d} already exists for this project"}
                fields["claim_number"] = new_no

        # Load current section_totals_json (dynamic sections)
        try:
            sec_totals = _json.loads(claim.get("section_totals_json") or "{}")
        except Exception:
            sec_totals = {}
        if not sec_totals:
            # backfill from legacy fixed columns
            sec_totals = {
                SECTION_CONSTRUCTION: {"label": "Construction Works",
                    "total": _safe_float(claim["section_a_total"]), "claimable": True},
                SECTION_PROVISIONAL_SUMS: {"label": "Provisional Sums",
                    "total": _safe_float(claim["section_b_total"]), "claimable": True},
                SECTION_PRELIMINARIES: {"label": "Preliminaries",
                    "total": _safe_float(claim["section_c_total"]), "claimable": True},
                SECTION_VARIATIONS: {"label": "Variations",
                    "total": _safe_float(claim["section_d_total"]), "claimable": True},
                SECTION_PS_EXCLUDED: {"label": "Provisional Sums Excluded",
                    "total": _safe_float(claim.get("section_e_total", 0)), "claimable": False},
            }

        # Merge edited section totals (section_totals: {code: total})
        if "section_totals" in fields and fields["section_totals"]:
            for code, val in fields["section_totals"].items():
                if code in sec_totals:
                    sec_totals[code]["total"] = _safe_float(val)
        # pop section_totals out so it isn't passed to update_claim
        fields.pop("section_totals", None)

        less_prev = _safe_float(fields.get("less_previous_claims",
                                           claim["less_previous_claims"]))
        retention = _safe_float(fields.get("retention_amount",
                                           claim["retention_amount"]))
        fields["less_previous_claims"] = less_prev
        fields["retention_amount"] = retention

        gross_claim = sum(_safe_float(v.get("total"))
                          for v in sec_totals.values() if v.get("claimable"))
        net_claim = gross_claim - less_prev - retention
        gst_amount = net_claim * 0.10
        total_incl_gst = net_claim + gst_amount

        db.update_claim(claim_entry_id, **fields)
        db.update_claim(claim_entry_id,
                        gross_claim=gross_claim,
                        cumulative_claimed=gross_claim,
                        less_previous_claims=less_prev,
                        retention_amount=retention,
                        total_retention_held=retention,
                        net_claim=net_claim,
                        gst_amount=gst_amount,
                        total_including_gst=total_incl_gst,
                        section_totals_json=_json.dumps(sec_totals))
        self.claim_updated.emit(claim_entry_id)
        return {"ok": True}

    def list_claims(self, project_entry_id: str) -> list[dict]:
        from shared_tools.progress_claim import progress_claim_db as db
        return db.get_claims(project_entry_id)

    def delete_claim(self, claim_entry_id: str) -> bool:
        """Delete a claim, its items, and any generated Excel/PDF files
        (wipe all traces)."""
        from shared_tools.progress_claim import progress_claim_db as db
        claim = db.get_claim(claim_entry_id)
        if not claim:
            return False
        for path_key in ("excel_path", "pdf_path"):
            p = claim.get(path_key)
            if p:
                try:
                    Path(p).unlink(missing_ok=True)
                except Exception:
                    pass
        ok = db.delete_claim(claim_entry_id)
        if ok:
            self.claim_updated.emit(claim_entry_id)
        return ok

    # ── Import a previously-finished claim ────────────────────────────

    def import_claim(self, project_entry_id: str, file_path: str) -> str | None:
        """Import a previously-finished claim from an XLSX or PDF file.

        Parses the file into a claim record (with per-item total_claimed,
        previously_claimed, current_claim), matches line items to the
        project's cashflow work items by description, and stores it. Future
        claims generated for later months will learn `previously_claimed`
        from this imported claim's per-item total_claimed.

        Returns the new claim entry_id, or None on failure.
        """
        p = Path(file_path)
        if not p.exists():
            raise FileNotFoundError(f"Claim file not found: {file_path}")
        suffix = p.suffix.lower()
        if suffix == ".xlsx":
            parsed = _parse_claim_xlsx(file_path)
        elif suffix == ".pdf":
            parsed = _parse_claim_pdf(file_path)
        else:
            raise ValueError(f"Unsupported file type: {suffix} (use .xlsx or .pdf)")
        if not parsed or not parsed.get("line_items"):
            raise ValueError("Could not extract any claim line items from the file.")
        return self._store_imported_claim(project_entry_id, parsed)

    def _store_imported_claim(self, project_entry_id: str, parsed: dict) -> str | None:
        from shared_tools.progress_claim import progress_claim_db as db

        project = db.get_project(project_entry_id)
        if not project:
            raise ValueError(f"Project not found: {project_entry_id}")

        work_items = db.get_work_items(project_entry_id)
        match_index = _build_description_index(work_items)

        claim_number = parsed.get("claim_number")
        if not claim_number:
            claim_number = db.next_claim_number(project_entry_id)
        else:
            # avoid collision with existing claim numbers
            existing_numbers = {c["claim_number"] for c in db.get_claims(project_entry_id)}
            base = claim_number
            while claim_number in existing_numbers:
                claim_number += 1
            claim_number = max(claim_number, base)

        claim_month = parsed.get("claim_month") or datetime.now().strftime("%Y-%m")
        claim_date = parsed.get("claim_date") or claim_month + "-01"
        entry_id = _new_entry_id()

        claim_items: list[dict] = []
        section_totals = {s: 0.0 for s in CLAIMABLE_SECTIONS}
        section_cumulative = {s: 0.0 for s in CLAIMABLE_SECTIONS}
        section_previous = {s: 0.0 for s in CLAIMABLE_SECTIONS}
        item_no_by_section = {s: 0 for s in CLAIMABLE_SECTIONS}
        sort_order = 0

        for li in parsed["line_items"]:
            desc = (li.get("description") or "").strip()
            if not desc:
                continue
            cost = _safe_float(li.get("cost"))
            total_claimed = _safe_float(li.get("total_claimed"))
            previously_claimed = _safe_float(li.get("previously_claimed"))
            current_claim = _safe_float(li.get("current_claim"))
            # Derive missing values to keep relationships consistent
            if total_claimed == 0 and current_claim != 0 and previously_claimed != 0:
                total_claimed = previously_claimed + current_claim
            if current_claim == 0:
                current_claim = total_claimed - previously_claimed
            if previously_claimed == 0 and total_claimed != 0 and current_claim != 0:
                previously_claimed = total_claimed - current_claim
            balance = cost - total_claimed if cost else 0.0
            cum_pct = (total_claimed / cost) if cost else 0.0

            # Match to a cashflow work item by description (exact then prefix)
            m = _match_work_item(desc, match_index, work_items)
            if m:
                wid, sec, item_cost = m
            else:
                wid, sec, item_cost = None, _infer_section(desc), None
            if cost == 0 and item_cost:
                cost = item_cost
                balance = cost - total_claimed
                cum_pct = (total_claimed / cost) if cost else 0.0

            if sec in item_no_by_section:
                item_no_by_section[sec] += 1
            sort_order += 1
            claim_items.append({
                "claim_entry_id": entry_id,
                "work_item_id": wid,
                "section": sec or SECTION_CONSTRUCTION,
                "item_number": item_no_by_section.get(sec, 0),
                "description": desc,
                "cost": cost,
                "cumulative_percentage": cum_pct,
                "total_claimed": total_claimed,
                "previously_claimed": previously_claimed,
                "current_claim": current_claim,
                "balance_remaining": balance,
                "sort_order": sort_order,
            })
            if sec in section_totals:
                section_totals[sec] += current_claim
                section_cumulative[sec] += total_claimed
                section_previous[sec] += previously_claimed

        gross_claim = sum(section_totals[s] for s in CLAIMABLE_SECTIONS)
        cumulative_claimed = sum(section_cumulative[s] for s in CLAIMABLE_SECTIONS)
        less_previous = sum(section_previous[s] for s in CLAIMABLE_SECTIONS)

        base_contract = _safe_float(project.get("base_contract_amount"))
        retention_cap = base_contract * 0.05 if base_contract > 0 else float("inf")
        total_retention_held = min(cumulative_claimed * 0.10, retention_cap)
        prior_retention = min(less_previous * 0.10, retention_cap)
        retention_amount = max(0.0, total_retention_held - prior_retention)
        net_claim = gross_claim - retention_amount
        gst_amount = net_claim * 0.10
        total_incl_gst = net_claim + gst_amount

        claim_data = {
            "entry_id": entry_id,
            "project_entry_id": project_entry_id,
            "claim_number": claim_number,
            "claim_month": claim_month,
            "claim_date": claim_date,
            "rev_number": 1,
            "status": "imported",
            "retention_percentage": 10,
            "retention_max_percentage": 5,
            "gross_claim": gross_claim,
            "less_previous_claims": less_previous,
            "retention_amount": retention_amount,
            "total_retention_held": total_retention_held,
            "net_claim": net_claim,
            "gst_amount": gst_amount,
            "total_including_gst": total_incl_gst,
            "section_a_total": section_totals[SECTION_CONSTRUCTION],
            "section_b_total": section_totals[SECTION_PROVISIONAL_SUMS],
            "section_c_total": section_totals[SECTION_PRELIMINARIES],
            "section_d_total": section_totals[SECTION_VARIATIONS],
            "section_e_total": section_totals.get(SECTION_PS_EXCLUDED, 0.0),
            "cumulative_claimed": cumulative_claimed,
        }
        db.upsert_claim(claim_data)
        db.clear_claim_items(entry_id)
        db.bulk_insert_claim_items(claim_items)
        self.claim_generated.emit(entry_id)
        return entry_id

    # ── Export ────────────────────────────────────────────────────────

    def export_excel(self, claim_entry_id: str) -> None:
        self._queue("export_excel", claim_entry_id=claim_entry_id)

    def _handle_export_excel(self, claim_entry_id: str) -> None:
        from shared_tools.progress_claim.progress_claim_template import build_claim_workbook
        from shared_tools.progress_claim import progress_claim_db as db
        summary = self.get_claim_summary(claim_entry_id)
        if not summary:
            raise ValueError(f"Claim not found: {claim_entry_id}")
        project = summary["project"] or {}
        safe_name = "".join(c if c.isalnum() or c in "-_ " else "_" for c in (project.get("name", "project")))[:60].strip() or "project"
        out_path = self._output_dir / f"{safe_name} - Claim {summary['claim']['claim_number']:02d} {summary['claim']['claim_month']}.xlsx"
        build_claim_workbook(summary, out_path)
        db.update_claim(claim_entry_id, excel_path=str(out_path))
        self.excel_generated.emit(claim_entry_id, str(out_path))

    def export_pdf(self, claim_entry_id: str) -> None:
        self._queue("export_pdf", claim_entry_id=claim_entry_id)

    def push_to_excel(self, project_entry_id: str) -> dict:
        """Push the current cashflow (from DB) back to the imported Excel file.

        The old file is backed up alongside as <stem>.backup_<ts>.xlsx, then
        the file is regenerated from the DB (sections, items, months, %/amount).
        Returns {"path": ..., "backup": ...} or {"error": ...}.
        """
        import shutil
        from shared_tools.progress_claim.progress_claim_template import build_cashflow_workbook
        from shared_tools.progress_claim import progress_claim_db as db
        project = db.get_project(project_entry_id)
        if not project:
            return {"error": "Project not found"}
        cashflow_path = project.get("cashflow_path")
        if not cashflow_path:
            return {"error": "No cashflow file path on this project (import one first)"}
        cashflow_path = Path(cashflow_path)
        if not cashflow_path.parent.exists():
            cashflow_path.parent.mkdir(parents=True, exist_ok=True)

        # Backup the old file
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = cashflow_path.parent / f"{cashflow_path.stem}.backup_{ts}.xlsx"
        if cashflow_path.exists():
            try:
                shutil.copy2(str(cashflow_path), str(backup_path))
            except Exception as e:
                return {"error": f"Backup failed: {e}"}

        state = self.get_cashflow(project_entry_id)
        try:
            build_cashflow_workbook(state, cashflow_path)
        except Exception as e:
            return {"error": f"Write failed: {e}"}
        self.progress_update.emit(100, "Cashflow pushed to Excel (old file backed up)")
        return {"path": str(cashflow_path), "backup": str(backup_path)}

    def _handle_export_pdf(self, claim_entry_id: str) -> None:
        from shared_tools.progress_claim import progress_claim_db as db
        claim = db.get_claim(claim_entry_id)
        if not claim:
            raise ValueError(f"Claim not found: {claim_entry_id}")
        excel_path = claim.get("excel_path")
        if not excel_path or not Path(excel_path).exists():
            # Generate Excel first
            self._handle_export_excel(claim_entry_id)
            claim = db.get_claim(claim_entry_id)
            excel_path = claim.get("excel_path")
        pdf_path = _excel_to_pdf(excel_path)
        db.update_claim(claim_entry_id, pdf_path=str(pdf_path))
        self.pdf_exported.emit(claim_entry_id, str(pdf_path))


# =============================================================================
# Cashflow parsing helpers
# =============================================================================


def _cell_by_letter(row, col_letter: str):
    """Get a cell value from an openpyxl row tuple by column letter."""
    from openpyxl.utils import column_index_from_string
    idx = column_index_from_string(col_letter) - 1
    return row[idx].value if idx < len(row) else None


def _classify_section_header(desc: str) -> str | None:
    """Return a section code if the row is a section header, else None."""
    d = desc.lower()
    if d.startswith("provisional sums") and "excluded" not in d and "reconciliation" not in d:
        return SECTION_PROVISIONAL_SUMS
    if d.startswith("overheads and preliminaries") or d.strip() == "preliminaries":
        return SECTION_PRELIMINARIES
    if d.startswith("client directed vari") or d.strip() == "variations":
        return SECTION_VARIATIONS
    if d.startswith("provisional sums excluded"):
        return SECTION_PS_EXCLUDED
    return None


def _is_skip_row(desc: str) -> bool:
    """Detect subtotal / total / margin / reconciliation / placeholder rows."""
    d = desc.lower().strip()
    skip_markers = (
        "total", "subtotal", "sub total", "reconciliation",
        "summary", "description", "cash flow", "cashflow", "arco project",
        "balance remain", "hollard perth", "% total", "construction total",
    )
    for m in skip_markers:
        if d.startswith(m) or d == m:
            return True
    return False


def _parse_cashflow_header(ws) -> dict:
    """Extract project metadata from the top rows of the cashflow sheet."""
    header = {}
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        if not row or row[0] is None:
            continue
        label = str(row[0]).strip().lower()
        val = row[1] if len(row) > 1 else None
        if not val:
            continue
        val_str = str(val).strip()
        if label.startswith("project") and "name" not in label:
            header["project_name"] = val_str
        elif label.startswith("client contact"):
            header["client_contact"] = val_str
        elif label.startswith("client"):
            header["client"] = val_str
        elif label.startswith("superintendent"):
            header["superintendent"] = val_str
        elif label.startswith("number"):
            header["job_number"] = val_str
        elif label.startswith("company"):
            header["company_name"] = val_str
    return header


def _detect_month_columns(ws) -> list[tuple[str, str, str]]:
    """Detect month columns from the header rows.

    Returns a list of (month_key, pct_col_letter, amt_col_letter) ordered by
    month index. Months are read from the date cells in the header row; the
    next row holds alternating '% Complete' / 'Amount' sub-headers.
    """
    from openpyxl.utils import get_column_letter

    months: list[tuple[str, str, str]] = []
    # Scan the first ~12 rows for the header containing date cells + a
    # '% Complete' / 'Amount' sub-header row.
    header_row_idx = None
    sub_row_idx = None
    for r in range(1, min(ws.max_row, 15) + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column, 40) + 1)]
        if any(isinstance(v, datetime) for v in row_vals):
            header_row_idx = r
        lowered = [str(v).strip().lower() if v is not None else "" for v in row_vals]
        if "% complete" in lowered and "amount" in lowered:
            sub_row_idx = r
            if header_row_idx is None:
                header_row_idx = r
            break

    if header_row_idx is None:
        return months

    # Determine which columns hold % vs Amount from the sub-header row.
    pct_cols: list[int] = []
    amt_cols: list[int] = []
    if sub_row_idx:
        for c in range(1, min(ws.max_column, 60) + 1):
            v = ws.cell(row=sub_row_idx, column=c).value
            if v is None:
                continue
            vs = str(v).strip().lower()
            if "complete" in vs:
                pct_cols.append(c)
            elif "amount" in vs:
                amt_cols.append(c)

    # If no explicit sub-header, fall back to alternating columns after 'Total'
    if not pct_cols:
        for c in range(3, min(ws.max_column, 60) + 1):
            v = ws.cell(row=header_row_idx, column=c).value
            if isinstance(v, datetime):
                pct_cols.append(c)
                amt_cols.append(c + 1)

    for c in pct_cols:
        date_val = ws.cell(row=header_row_idx, column=c).value
        mkey = _month_key_from_date(date_val)
        if not mkey:
            # Some layouts put the date one column earlier; try c-1
            alt = ws.cell(row=header_row_idx, column=c - 1).value
            mkey = _month_key_from_date(alt)
        if not mkey:
            continue
        amt_c = c + 1
        if amt_c not in amt_cols and amt_cols:
            # find the matching amount column to the right
            amt_c = next((a for a in amt_cols if a > c), c + 1)
        months.append((mkey, get_column_letter(c), get_column_letter(amt_c)))

    return months


# =============================================================================
# Claim import helpers (XLSX / PDF parsing + description matching)
# =============================================================================


def _norm_desc(s: str) -> str:
    """Normalize a description for fuzzy matching."""
    import re
    s = str(s or "").lower()
    s = re.sub(r"\s+", " ", s).strip()
    s = re.sub(r"[^a-z0-9 ]", "", s)
    return s.strip()


def _build_description_index(work_items: list[dict]) -> dict[str, tuple]:
    """Map normalized description -> (work_item_id, section, cost)."""
    idx: dict[str, tuple] = {}
    for it in work_items:
        idx[_norm_desc(it["description"])] = (it["id"], it["section"], it["cost"])
    return idx


def _infer_section(desc: str) -> str:
    d = desc.lower()
    # "margin" before "variation" — "Margins (incl. Variations)" is a
    # preliminaries margin, not a variation.
    if any(k in d for k in ("insurance", "prelim", "staffing", "petty", "clean",
                            "site ", "fencing", "delapidation", "margin")):
        return SECTION_PRELIMINARIES
    if d.startswith("vo") or "variation" in d:
        return SECTION_VARIATIONS
    if "provisional sum" in d and "exclud" in d:
        return SECTION_PS_EXCLUDED
    return SECTION_CONSTRUCTION


def _match_work_item(desc: str, match_index: dict[str, tuple],
                     work_items: list[dict]) -> tuple | None:
    """Match a parsed line description to a cashflow work item.

    Tries exact normalized match, then prefix match (one description starts
    with the other), so e.g. "Margins (incl. Variations)" → "Margins".
    Returns (work_item_id, section, cost) or None.
    """
    nd = _norm_desc(desc)
    if nd in match_index:
        return match_index[nd]
    # prefix match
    for it in work_items:
        nd2 = _norm_desc(it["description"])
        if not nd2:
            continue
        if nd.startswith(nd2) or nd2.startswith(nd):
            return (it["id"], it["section"], it["cost"])
    return None


def _parse_claim_number(text: str) -> int | None:
    import re
    m = re.search(r"(?:claim\s*no\.?\s*|claim\s*#?\s*|pc\s*0*)(\d{1,3})", str(text or "").lower())
    if m:
        return int(m.group(1))
    return None


def _parse_claim_month_from_text(text: str) -> str | None:
    """Try to find a YYYY-MM or date in text; fall back to month name."""
    import re
    s = str(text or "")
    m = re.search(r"(20\d{2})[-/](\d{1,2})", s)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}"
    m = re.search(r"(20\d{2})", s)
    months = {"jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
              "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12}
    mm = re.search(r"\b(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*\b", s.lower())
    if m and mm:
        return f"{m.group(1)}-{months[mm.group(1)]:02d}"
    return None


def _parse_claim_xlsx(xlsx_path: str) -> dict | None:
    """Parse an exported/external progress claim xlsx into a claim dict.

    Looks for the detail sheet (columns matching Description / TOTAL CLAIMED /
    PREVIOUSLY CLAIMED / CURRENT CLAIM) and the claim number/month from any
    sheet. Returns {claim_number, claim_month, claim_date, line_items:[...]}.
    """
    import openpyxl
    import re
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    claim_number = None
    claim_month = None
    claim_date = None
    line_items: list[dict] = []

    # Header hints that identify a claim-detail sheet
    detail_keywords = ("total claimed", "previously claimed", "current claim", "balance")

    def _scan_meta(ws):
        """Scan one sheet for claim number / month / date meta."""
        nonlocal claim_number, claim_month, claim_date
        for r in range(1, min(ws.max_row, 30) + 1):
            for c in range(1, min(ws.max_column, 12) + 1):
                v = ws.cell(row=r, column=c).value
                if v is None:
                    continue
                s = str(v)
                if claim_number is None:
                    cn = _parse_claim_number(s)
                    if cn:
                        claim_number = cn
                if isinstance(v, datetime):
                    if claim_month is None:
                        claim_month = f"{v.year:04d}-{v.month:02d}"
                    if claim_date is None:
                        claim_date = v.strftime("%Y-%m-%d")
                elif claim_month is None:
                    cm = _parse_claim_month_from_text(s)
                    if cm:
                        claim_month = cm

    def _find_detail(ws):
        """Find the detail header row + column map on a sheet, or None."""
        for r in range(1, min(ws.max_row, 40) + 1):
            cells = [ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column, 20) + 1)]
            lowered = [re.sub(r"\s+", " ", str(v).strip().lower()) if v is not None else "" for v in cells]
            joined = " | ".join(lowered)
            if any(k in joined for k in detail_keywords):
                col_map: dict[str, int] = {}
                for ci, lab in enumerate(lowered):
                    if "description" in lab or "trade" in lab:
                        col_map["description"] = ci
                    elif "cost" in lab:
                        col_map["cost"] = ci
                    elif "complete" in lab and "%" in lab:
                        col_map["cum_pct"] = ci
                    elif "complete" in lab:
                        col_map["cum_pct"] = ci
                    elif "total claimed" in lab:
                        col_map["total"] = ci
                    elif "previously" in lab:
                        col_map["previously"] = ci
                    elif "current" in lab:
                        col_map["current"] = ci
                    elif "balance" in lab:
                        col_map["balance"] = ci
                if "description" in col_map:
                    return r, col_map
        return None, None

    # Meta: scan ALL sheets (the date usually lives on the summary sheet).
    for ws in wb.worksheets:
        _scan_meta(ws)

    # Detail: prefer named sheets, then any sheet with a detail header.
    detail_ws = None
    header_row = col_map = None
    for name in ("Detail", "Sheet3"):
        if name in wb.sheetnames:
            hr, cm = _find_detail(wb[name])
            if hr is not None:
                detail_ws, header_row, col_map = wb[name], hr, cm
                break
    if detail_ws is None:
        for ws in wb.worksheets:
            hr, cm = _find_detail(ws)
            if hr is not None:
                detail_ws, header_row, col_map = ws, hr, cm
                break

    if header_row and col_map and "description" in col_map:
        section = SECTION_CONSTRUCTION
        for r in range(header_row + 1, detail_ws.max_row + 1):
            desc_v = detail_ws.cell(row=r, column=col_map["description"] + 1).value
            if desc_v is None:
                continue
            desc = str(desc_v).strip()
            if not desc:
                continue
            # section headers
            sec = _classify_section_header(desc)
            if sec is not None:
                section = sec
                continue
            if _is_skip_row(desc):
                continue
            # claim-parser extra skips: subtotal / contract-sum / grand-total
            # rows that don't start with "subtotal"/"total" (e.g. "MAIN WORKS
            # CONTRACT SUM SUBTOTAL (A+B)").
            dl = desc.lower()
            if "subtotal" in dl or "contract sum" in dl or "balance remain" in dl:
                continue
            def g(key):
                if key not in col_map:
                    return 0.0
                return _safe_float(detail_ws.cell(row=r, column=col_map[key] + 1).value)
            cost = g("cost"); total = g("total"); prev = g("previously")
            curr = g("current"); bal = g("balance"); cum = g("cum_pct")
            # need at least some monetary value to be a real line
            if total == 0 and curr == 0 and prev == 0 and cost == 0:
                continue
            # validity: a real line item has current ≤ total ≤ cost and
            # previously ≤ total. Summary/annotation rows in the bottom block
            # (e.g. "Approved", "PS in discussion") leak large values into the
            # current column while total is 0 — drop those.
            if total > 0:
                if curr > total + 1 or prev > total + 1:
                    continue
            else:
                if curr > 0 or prev > 0:
                    continue
            line_items.append({
                "description": desc,
                "cost": cost,
                "cumulative_percent": cum,
                "total_claimed": total,
                "previously_claimed": prev,
                "current_claim": curr,
                "balance_remaining": bal,
                "section": section,
            })

    wb.close()
    if not line_items:
        return None
    return {
        "claim_number": claim_number,
        "claim_month": claim_month,
        "claim_date": claim_date,
        "line_items": line_items,
    }


def _parse_claim_pdf(pdf_path: str) -> dict | None:
    """Parse a progress claim PDF: extract text (pdfplumber/fitz), then ask
    the LLM to return a structured JSON table. Falls back to Gemini vision
    (page images) if little text is extractable."""
    text = _extract_pdf_text(pdf_path)
    parsed = None
    if text and len(text) > 300:
        parsed = _llm_extract_claim(text)
    if not parsed or not parsed.get("line_items"):
        parsed = _llm_extract_claim_vision(pdf_path)
    return parsed


def _extract_pdf_text(pdf_path: str, max_pages: int = 12) -> str:
    texts: list[str] = []
    try:
        import pdfplumber
        with pdfplumber.open(str(pdf_path)) as pdf:
            for i, page in enumerate(pdf.pages):
                if i >= max_pages:
                    break
                t = page.extract_text() or ""
                if t:
                    texts.append(t)
    except Exception:
        pass
    if texts:
        return "\n".join(texts).strip()
    try:
        import fitz  # type: ignore
        doc = fitz.open(str(pdf_path))
        for i in range(min(doc.page_count, max_pages)):
            texts.append(doc[i].get_text())
        doc.close()
    except Exception:
        pass
    return "\n".join(texts).strip()


_CLAIM_EXTRACT_PROMPT = """You are a construction progress-claim analyst. Extract the structured claim data from the text below.

Return ONLY a JSON object (no markdown, no prose) with this shape:
{
  "claim_number": <integer or null>,
  "claim_month": "YYYY-MM" or null,
  "claim_date": "YYYY-MM-DD" or null,
  "line_items": [
    {"description": "<text>", "cost": <number>, "cumulative_percent": <number 0-100>,
     "total_claimed": <number>, "previously_claimed": <number>,
     "current_claim": <number>, "balance_remaining": <number>}
  ]
}

Rules:
- cumulative_percent is a percentage 0-100 (not a fraction).
- Only include real work line items (skip section headers like "CONSTRUCTION WORKS", subtotals, totals, "Balance remain").
- If a field is not present in the source, use 0 for numbers and null for claim_number/date.
- Keep descriptions concise but faithful."""


def _llm_extract_claim(text: str) -> dict | None:
    from shared_tools.core.llm_config import get_llm
    try:
        llm = get_llm("smart")
        result = llm.call(_CLAIM_EXTRACT_PROMPT + "\n\n--- CLAIM TEXT ---\n" + text[:12000])
        raw = result.strip() if isinstance(result, str) else str(result)
        return _parse_json_obj(raw)
    except Exception as e:
        print(f"[progress_claim] LLM claim extraction failed: {e}")
        return None


def _llm_extract_claim_vision(pdf_path: str) -> dict | None:
    """Fallback: render PDF pages to PNGs and send to Gemini vision."""
    import os
    try:
        import fitz  # type: ignore
        import google.generativeai as genai  # type: ignore
    except Exception as e:
        print(f"[progress_claim] vision deps unavailable: {e}")
        return None
    api_key = os.environ.get("GEMINI_API_KEY")
    if not api_key:
        return None
    try:
        genai.configure(api_key=api_key)
        model_name = os.environ.get("GEMINI_VISION_MODEL", "gemini-2.5-flash")
        model = genai.GenerativeModel(model_name)
        doc = fitz.open(str(pdf_path))
        parts: list = [_CLAIM_EXTRACT_PROMPT]
        for i in range(min(doc.page_count, 10)):
            pix = doc[i].get_pixmap(dpi=120)
            parts.append({"mime_type": "image/png", "data": pix.tobytes("png")})
        doc.close()
        resp = model.generate_content(parts)
        raw = (resp.text or "").strip()
        return _parse_json_obj(raw)
    except Exception as e:
        print(f"[progress_claim] vision claim extraction failed: {e}")
        return None


def _parse_json_obj(text: str) -> dict | None:
    import json, re
    if not text:
        return None
    try:
        return json.loads(text.strip())
    except Exception:
        pass
    m = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", text, re.DOTALL)
    if m:
        try:
            return json.loads(m.group(1))
        except Exception:
            pass
    m = re.search(r"\{.*\}", text, re.DOTALL)
    if m:
        try:
            return json.loads(m.group(0))
        except Exception:
            pass
    return None


# =============================================================================
# Excel → PDF (COM)
# =============================================================================


def _excel_to_pdf(xlsx_path: str) -> str:
    """Convert an xlsx to PDF via Excel COM (Windows). Falls back to a copy
    of the xlsx if COM is unavailable."""
    pdf_path = str(Path(xlsx_path).with_suffix(".pdf"))
    try:
        import win32com.client  # type: ignore
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        try:
            wb = excel.Workbooks.Open(os.path.abspath(xlsx_path))
            # Fit-to-width for clean PDF
            for ws in wb.Worksheets:
                ws.PageSetup.FitToPagesWide = 1
                ws.PageSetup.FitToPagesTall = False
                ws.PageSetup.Zoom = False
            wb.ExportAsFixedFormat(0, os.path.abspath(pdf_path))
            wb.Close(False)
        finally:
            excel.Quit()
        return pdf_path
    except Exception as e:
        print(f"[progress_claim] Excel COM PDF export failed ({e}); returning xlsx path")
        return xlsx_path
