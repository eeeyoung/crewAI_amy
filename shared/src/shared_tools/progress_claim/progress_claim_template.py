"""Progress Claim Excel/PDF builder.

Builds the claim workbook from scratch, strictly following Sheet1 of the source
example ("Econolodge Progress Claim Advice 6 - 2598 May26.3.xlsx"):

  Sheet "Sheet1"  — the PROJECT PAYMENT ADVICE cover page:
                     ABN/company header, claim meta, the bordered section
                     table (A CONSTRUCTION WORKS / B PROVISIONAL SUMS /
                     C PRELIMINARIES / D OPTIONS / D VARIATIONS), and the
                     summary block (Gross cumulative, Less previous after
                     retention, Retention total held, Net, GST, Total incl GST)
                     with a right-hand incl-GST / balance-remain / max-retention
                     column.
  Sheet "Detail"  — the line-item breakdown grouped by section.

Styling: Arial throughout, pure black & white (no fills, black borders), to
match the requested template. All figures come pre-computed from
ProgressClaimService; this module only renders them.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from shared_tools.progress_claim.progress_claim_service import (
    CLAIMABLE_SECTIONS,
    SECTION_LABELS,
)

# ── Style constants (Arial, black & white) ─────────────────────────────────

_ARIAL = "Arial"
_BLACK = "FF000000"

_MEDIUM = Side(style="medium", color=_BLACK)
_THIN = Side(style="thin", color=_BLACK)

_BOX_MED = Border(left=_MEDIUM, right=_MEDIUM, top=_MEDIUM, bottom=_MEDIUM)

def _font(size=10, bold=False, italic=False):
    return Font(name=_ARIAL, size=size, bold=bold, italic=italic, color=_BLACK)

_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
_LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
_RIGHT = Alignment(horizontal="right", vertical="center")
_CENTER = Alignment(horizontal="center", vertical="center")

# Accounting currency format (matches the source template's money cells)
_MONEY_FMT = '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'
_PCT_FMT = '0%'


def _money(cell, value):
    cell.value = value
    cell.number_format = _MONEY_FMT
    cell.alignment = _RIGHT
    cell.font = _font()


def _pct(cell, value):
    cell.value = value
    cell.number_format = _PCT_FMT
    cell.alignment = _CENTER
    cell.font = _font()


def _label(cell, value, bold=False, align=_LEFT, size=10):
    cell.value = value
    cell.font = _font(size=size, bold=bold)
    cell.alignment = align


# =============================================================================
# Public API
# =============================================================================


def build_claim_workbook(summary: dict, output_path: Path) -> Path:
    """Build the full claim workbook (Sheet1 + Detail)."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    _build_sheet1(wb.active, summary)
    detail = wb.create_sheet("Detail")
    _build_detail_sheet(detail, summary)

    wb.save(output_path)
    return output_path


# =============================================================================
# Cashflow workbook (push-to-Excel: regenerate the imported cashflow from DB)
# =============================================================================


def build_cashflow_workbook(state: dict, output_path: Path) -> Path:
    """Build a cashflow xlsx from the DB cashflow state (project, sections,
    items, months, progress), modeled on the source cashflow layout.

    state: dict from ProgressClaimService.get_cashflow().
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    project = state.get("project") or {}
    section_defs = state.get("section_defs") or []
    sections = state.get("sections") or {}
    months = state.get("months") or []

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 14
    for i in range(len(months)):
        ws.column_dimensions[get_column_letter(3 + i * 2)].width = 9      # %
        ws.column_dimensions[get_column_letter(4 + i * 2)].width = 13     # $

    money_fmt = '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'

    # Header
    r = 1
    _set(ws, r, 1, "Project"); _set(ws, r, 2, project.get("name", ""), bold=True); r += 1
    _set(ws, r, 1, "Number"); _set(ws, r, 2, project.get("job_number", "")); r += 1
    _set(ws, r, 1, "Client"); _set(ws, r, 2, project.get("client", "")); r += 1
    _set(ws, r, 1, "Client contact"); _set(ws, r, 2, project.get("client_contact", "")); r += 1
    _set(ws, r, 1, "Superintendent"); _set(ws, r, 2, project.get("superintendent", "")); r += 1
    _set(ws, r, 1, "CASH FLOW STATEMENT", bold=True); r += 1
    r += 1

    # Month header row
    _set(ws, r, 1, "Description", bold=True)
    _set(ws, r, 2, "Total", bold=True)
    for i, m in enumerate(months):
        _set(ws, r, 3 + i * 2, m.get("month_label", ""), bold=True, align="center")
        _set(ws, r, 4 + i * 2, "", bold=True, align="center")
    month_header_row = r
    r += 1
    # % / Amount sub-header
    _set(ws, r, 1, "", )
    _set(ws, r, 2, "")
    for i in range(len(months)):
        _set(ws, r, 3 + i * 2, "% Complete", bold=False, align="center")
        _set(ws, r, 4 + i * 2, "Amount", bold=False, align="center")
    r += 1

    # Sections + items
    for sd in section_defs:
        code = sd["section_code"]
        label = sd.get("section_label") or code
        items = sections.get(code, [])
        _set(ws, r, 1, f"{label}", bold=True)
        _box(ws, r, 1, r, 2 + len(months) * 2, medium=False)
        r += 1
        for it in items:
            is_margin = it.get("item_type") == "margin"
            _set(ws, r, 1, it.get("description", ""))
            cost = it.get("cost", 0)
            c = ws.cell(row=r, column=2, value=cost)
            c.number_format = money_fmt; c.font = _font()
            for i, m in enumerate(months):
                p = next((g for g in it.get("progress", []) if g.get("month_id") == m["id"]), None)
                pct = p["percentage"] if p else 0
                amt = p["amount"] if p else 0
                if is_margin:
                    # Margin: leave % cell blank
                    pc = ws.cell(row=r, column=3 + i * 2, value="")
                    pc.font = _font(); pc.alignment = Alignment(horizontal="center")
                else:
                    pc = ws.cell(row=r, column=3 + i * 2, value=pct)
                    pc.number_format = "0.00%"; pc.font = _font(); pc.alignment = Alignment(horizontal="right")
                ac = ws.cell(row=r, column=4 + i * 2, value=amt)
                ac.number_format = money_fmt; ac.font = _font(); ac.alignment = Alignment(horizontal="right")
            r += 1

    ws.freeze_panes = "C{}".format(month_header_row + 2)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    wb.save(output_path)
    return output_path


def _set(ws, r, c, value, bold=False, align=None):
    cell = ws.cell(row=r, column=c, value=value)
    cell.font = _font(bold=bold)
    if align:
        cell.alignment = Alignment(horizontal=align)
    return cell


# =============================================================================
# Sheet1 — PROJECT PAYMENT ADVICE (strict template reproduction)
# =============================================================================

# Section rows in template order: (item_letter, label, source_field)
_SHEET1_SECTIONS = [
    ("A", "CONSTRUCTION WORKS", "section_a_total"),
    ("B", "PROVISIONAL SUMS", "section_b_total"),
    ("C", "PRELIMINARIES", "section_c_total"),
    ("D", "OPTIONS", "section_e_total"),        # PS Excluded shown as OPTIONS
    ("D", "VARIATIONS", "section_d_total"),
]


def _build_sheet1(ws, summary: dict) -> None:
    claim = summary["claim"]
    project = summary.get("project") or {}

    ws.title = "Sheet1"
    ws.sheet_view.showGridLines = False

    # Column widths (match source: A margin, B item, C description, D period,
    # E total ex-GST, F right incl-GST/balance, G incl-GST label)
    widths = {"A": 8.14, "B": 5.43, "C": 48.29, "D": 14.71,
              "E": 26.14, "F": 17.71, "G": 14.0}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    base_contract = _f(project.get("base_contract_amount"))
    cumulative_claimed = _f(claim.get("cumulative_claimed"))
    # Under the cumulative model, less_previous_claims is already
    # "Less Previous (after retention)" and retention_amount is the total
    # held to date — both editable on the summary card.
    less_previous_after_retention = _f(claim.get("less_previous_claims"))
    retention_amount = _f(claim.get("retention_amount"))
    net_claim = _f(claim.get("net_claim"))
    gst_amount = _f(claim.get("gst_amount"))
    total_incl_gst = _f(claim.get("total_including_gst"))
    retention_max_pct = _f(claim.get("retention_max_percentage"), 5.0) or 5.0

    max_retention = base_contract * retention_max_pct / 100.0
    balance_remaining = base_contract - cumulative_claimed

    claim_no = claim.get("claim_number", 1)
    rev = claim.get("rev_number", 1)
    month_label = _month_upper(claim.get("claim_month", ""))

    # ── Header block ───────────────────────────────────────────────────
    r = 7
    _label(ws.cell(row=r, column=3), f"ABN. No. {project.get('company_abn', '') or '____________'}", bold=True)
    r = 8
    ws.row_dimensions[r].height = 30
    _label(ws.cell(row=r, column=3), "PROJECT PAYMENT ADVICE", bold=True, size=16)
    r = 9
    _label(ws.cell(row=r, column=4), "From:", bold=True, align=_RIGHT)
    _label(ws.cell(row=r, column=5), project.get("company_name") or "Welink Construction Pty Ltd")
    r = 10
    _label(ws.cell(row=r, column=5), project.get("company_address") or "")
    r = 11
    _label(ws.cell(row=r, column=5), "")  # address line 2 / suburb

    # ── Claim meta ─────────────────────────────────────────────────────
    r = 16
    _label(ws.cell(row=r, column=3), f"Progress Claim No.{claim_no:02d}", bold=True, size=11)
    _label(ws.cell(row=r, column=4), "Date:", bold=True, align=_RIGHT)
    _label(ws.cell(row=r, column=5), _fmt_date(claim.get("claim_date")))
    r = 17
    _label(ws.cell(row=r, column=4), "Rev:", bold=True, align=_RIGHT)
    _label(ws.cell(row=r, column=5), rev)
    r = 18
    _label(ws.cell(row=r, column=2), "Project:", bold=True, align=_RIGHT)
    proj_desc = project.get("name", "")
    if project.get("site_address"):
        proj_desc = f"{proj_desc} - {project['site_address']}" if proj_desc else project["site_address"]
    _label(ws.cell(row=r, column=3), proj_desc)
    r = 19
    ws.row_dimensions[r].height = 20
    _label(ws.cell(row=r, column=2), "Job No:", bold=True, align=_RIGHT)
    _label(ws.cell(row=r, column=3), project.get("job_number", ""))
    _label(ws.cell(row=r, column=4), "Advice payment due", bold=True, align=_CENTER)
    _label(ws.cell(row=r, column=5), _fmt_date(claim.get("claim_date")))
    r = 21
    _label(ws.cell(row=r, column=2), "Contract No:", bold=True, align=_RIGHT)
    _label(ws.cell(row=r, column=3), project.get("job_number", ""))
    r = 22
    _label(ws.cell(row=r, column=3), " ")

    # ── Section table header (row 23) + period (row 24) ────────────────
    hdr_r = 23
    _label(ws.cell(row=hdr_r, column=2), "Item", bold=True)
    _label(ws.cell(row=hdr_r, column=3), "Description", bold=True)
    _label(ws.cell(row=hdr_r, column=4), "For Period", bold=True, align=_CENTER)
    _label(ws.cell(row=hdr_r, column=5), "Total", bold=True, align=_CENTER)
    _box(ws, hdr_r, 2, hdr_r, 5, medium=True)

    _label(ws.cell(row=24, column=4), month_label, bold=True, align=_CENTER)
    _box(ws, 24, 2, 24, 5, medium=True)

    # Section rows at 26, 29, 32, 35, 38
    section_rows = [26, 29, 32, 35, 38]
    for (letter, label, field), rr in zip(_SHEET1_SECTIONS, section_rows):
        _label(ws.cell(row=rr, column=2), letter, bold=True)
        _label(ws.cell(row=rr, column=3), label, bold=True)
        _money(ws.cell(row=rr, column=5), _f(claim.get(field)))
    # Section table body borders (rows 25-40, cols B-E):
    # outer left/right medium, B|C thin, C|D & D|E medium, bottom medium.
    _box(ws, 25, 2, 40, 5, medium=True)
    _vdiv(ws, 25, 40, 2, medium=False)   # B|C thin
    _vdiv(ws, 25, 40, 3, medium=True)    # C|D medium
    _vdiv(ws, 25, 40, 4, medium=True)    # D|E medium
    _label(ws.cell(row=40, column=6), "Balance remain", align=_RIGHT)

    # ── Summary block (rows 41-51) ─────────────────────────────────────
    _label(ws.cell(row=41, column=3), "Gross Claim for Works Completed ")
    _money(ws.cell(row=41, column=5), cumulative_claimed)
    _money(ws.cell(row=41, column=6), balance_remaining)

    _label(ws.cell(row=42, column=3), "Less Previous Progress Claims to Date (after retention)")
    _money(ws.cell(row=42, column=5), less_previous_after_retention)
    _label(ws.cell(row=42, column=6), "Max retention", align=_RIGHT)

    _label(ws.cell(row=43, column=3), "Retention Amount (10% to a maximum of 5%)")
    _money(ws.cell(row=43, column=5), retention_amount)
    _money(ws.cell(row=43, column=6), max_retention)

    _label(ws.cell(row=44, column=3), "Net Amount Claimed", bold=True)
    _money(ws.cell(row=44, column=5), net_claim)
    _money(ws.cell(row=44, column=6), net_claim * 1.1)
    _label(ws.cell(row=44, column=7), "incl GST", align=_LEFT)

    # Adjustment line (row 45) — kept blank/0 to match template
    _money(ws.cell(row=45, column=5), 0)
    _money(ws.cell(row=45, column=6), 0)
    _label(ws.cell(row=45, column=7), "incl GST", align=_LEFT)

    _label(ws.cell(row=46, column=3), "This Claim", bold=True)
    _money(ws.cell(row=46, column=5), net_claim)
    _money(ws.cell(row=46, column=6), 0)

    _label(ws.cell(row=49, column=3), "Add GST ")
    _pct(ws.cell(row=49, column=4), 0.1)
    _money(ws.cell(row=49, column=5), gst_amount)

    _label(ws.cell(row=51, column=3), "Total Amount Claimed Including GST", bold=True)
    _money(ws.cell(row=51, column=5), total_incl_gst)
    _money(ws.cell(row=51, column=6), total_incl_gst)
    _label(ws.cell(row=51, column=7), "incl GST", align=_LEFT)

    # Borders for the summary block (rows 41-51, cols B-E):
    # outer medium box, B|C thin divider, thin separators between rows.
    _box(ws, 41, 2, 51, 5, medium=True)
    _vdiv(ws, 41, 51, 2, medium=False)  # B|C thin
    for rr in [41, 42, 43, 44, 45, 46, 49, 50]:
        for cc in range(2, 6):
            _side(ws, rr, cc, bottom=_THIN)

    # Print setup: fit to one page wide
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    try:
        ws.sheet_properties.pageSetUpPr.fitToPage = True
    except Exception:
        pass


# =============================================================================
# Detail sheet
# =============================================================================


def _build_detail_sheet(ws, summary: dict) -> None:
    claim = summary["claim"]
    project = summary.get("project") or {}
    items = summary.get("items", [])

    ws.sheet_view.showGridLines = False
    widths = {"A": 6, "B": 52, "C": 14, "D": 14, "E": 12, "F": 16, "G": 16, "H": 16, "I": 16}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    _label(ws.cell(row=1, column=2), f"Progress Claim No.{claim.get('claim_number', 1):02d}", bold=True, size=16)
    _label(ws.cell(row=2, column=2), project.get("name", ""), bold=True)
    _label(ws.cell(row=3, column=2),
           f"For Period: {claim.get('claim_month', '')}    Date: {_fmt_date(claim.get('claim_date'))}", size=9)

    r = 5
    headers = ["Item", "TRADES", "COST", "% COMPLETE", "TOTAL CLAIMED",
               "PREVIOUSLY CLAIMED", "CURRENT CLAIM", "Balance Remaining"]
    for i, h in enumerate(headers):
        c = ws.cell(row=r, column=1 + i)
        _label(c, h, bold=True, align=_CENTER)
    _box(ws, r, 1, r, 8, medium=True)
    ws.freeze_panes = "A6"
    r += 1

    sections = summary.get("sections", {})
    order = [s for s in CLAIMABLE_SECTIONS if sections.get(s)] + \
            [s for s in sections if s not in CLAIMABLE_SECTIONS]

    for sec in order:
        sec_items = sections.get(sec, [])
        if not sec_items:
            continue
        label = SECTION_LABELS.get(sec, sec)
        _label(ws.cell(row=r, column=1), sec, bold=True)
        _label(ws.cell(row=r, column=2), label, bold=True)
        for cc in range(1, 9):
            _side(ws, r, cc, bottom=_MEDIUM)
        r += 1

        sec_cost = sec_total = sec_prev = sec_current = sec_balance = 0.0
        for it in sec_items:
            _label(ws.cell(row=r, column=1), it.get("item_number", ""), align=_CENTER)
            _label(ws.cell(row=r, column=2), it.get("description", ""), align=_LEFT_TOP)
            _money(ws.cell(row=r, column=3), it.get("cost", 0))
            cum_pct = it.get("cumulative_percentage", 0)
            if cum_pct == 0:
                _label(ws.cell(row=r, column=4), "—", align=_CENTER)
            else:
                _pct(ws.cell(row=r, column=4), cum_pct)
            _money(ws.cell(row=r, column=5), it.get("total_claimed", 0))
            _money(ws.cell(row=r, column=6), it.get("previously_claimed", 0))
            c = ws.cell(row=r, column=7); _money(c, it.get("current_claim", 0)); c.font = _font(bold=True)
            _money(ws.cell(row=r, column=8), it.get("balance_remaining", 0))
            for cc in range(1, 9):
                ws.cell(row=r, column=cc).border = Border(
                    left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
            sec_cost += _f(it.get("cost")); sec_total += _f(it.get("total_claimed"))
            sec_prev += _f(it.get("previously_claimed")); sec_current += _f(it.get("current_claim"))
            sec_balance += _f(it.get("balance_remaining"))
            r += 1

        _label(ws.cell(row=r, column=2), f"SUBTOTAL {sec}", bold=True)
        for col, val in [(3, sec_cost), (5, sec_total), (6, sec_prev), (7, sec_current), (8, sec_balance)]:
            c = ws.cell(row=r, column=col); _money(c, val); c.font = _font(bold=True)
        _box(ws, r, 1, r, 8, medium=True)
        r += 2

    _label(ws.cell(row=r, column=2), "TOTAL (A+B+C+D)", bold=True, size=12)
    totals = {
        3: sum(_f(it.get("cost")) for it in items if it.get("section") in CLAIMABLE_SECTIONS),
        5: _f(claim.get("cumulative_claimed")),
        6: _f(claim.get("less_previous_claims")),
        7: _f(claim.get("gross_claim")),
    }
    for col, val in totals.items():
        c = ws.cell(row=r, column=col); _money(c, val); c.font = _font(bold=True, size=12)
    _box(ws, r, 1, r, 8, medium=True)
    r += 2
    _label(ws.cell(row=r, column=2), "Net Amount Claimed", bold=True)
    _money(ws.cell(row=r, column=7), _f(claim.get("net_claim")))
    r += 1
    _label(ws.cell(row=r, column=2), "Total Including GST", bold=True, size=12)
    _money(ws.cell(row=r, column=7), _f(claim.get("total_including_gst")))

    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True


# =============================================================================
# Helpers
# =============================================================================


def _f(v, default=0.0):
    try:
        if v is None:
            return default
        x = float(v)
        return x if x == x else default  # NaN guard
    except (TypeError, ValueError):
        return default


def _fmt_date(v):
    if not v:
        return ""
    s = str(v)[:10]
    # accept YYYY-MM-DD
    try:
        from datetime import datetime
        d = datetime.strptime(s, "%Y-%m-%d")
        return d.strftime("%d-%b-%y")
    except Exception:
        return s


def _month_upper(month_key: str) -> str:
    months = ["JAN", "FEB", "MAR", "APR", "MAY", "JUN",
              "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]
    try:
        _y, m = str(month_key).split("-")
        return months[int(m) - 1]
    except Exception:
        return str(month_key).upper()


def _side(ws, r, c, *, left=None, right=None, top=None, bottom=None):
    """Set individual border sides on a cell, preserving unspecified sides."""
    cell = ws.cell(row=r, column=c)
    b = cell.border
    cell.border = Border(
        left=left or b.left,
        right=right or b.right,
        top=top or b.top,
        bottom=bottom or b.bottom,
    )


def _box(ws, r1, c1, r2, c2, medium=True):
    side = _MEDIUM if medium else _THIN
    for rr in range(r1, r2 + 1):
        _side(ws, rr, c1, left=side)
        _side(ws, rr, c2, right=side)
    for cc in range(c1, c2 + 1):
        _side(ws, r1, cc, top=side)
        _side(ws, r2, cc, bottom=side)


def _vdiv(ws, r1, r2, col, medium=False):
    """Vertical divider on the RIGHT edge of `col`, rows r1..r2."""
    side = _MEDIUM if medium else _THIN
    for rr in range(r1, r2 + 1):
        _side(ws, rr, col, right=side)


def _hdiv(ws, r1, r2, c1, c2, medium=False):
    """Horizontal divider on the BOTTOM edge of row r1, cols c1..c2
    (applied to every row r1..r2-1 as the bottom of that row)."""
    side = _MEDIUM if medium else _THIN
    for rr in range(r1, r2):
        for cc in range(c1, c2 + 1):
            _side(ws, rr, cc, bottom=side)
